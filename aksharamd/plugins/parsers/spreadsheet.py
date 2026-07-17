from __future__ import annotations

import csv as _csv
import io
from pathlib import Path
from typing import TYPE_CHECKING

from ...context import CompilationContext
from ...models.block import Block, BlockType, ExtractionConfidence
from ...models.document import Document
from ...models.table import ExtractionMethod, TableCell, TableData
from ..base import ParserPlugin
from ..registry import register_parser

if TYPE_CHECKING:
    from ...models.key_value import KeyValueGroup

_MAX_ROWS_PER_SHEET = 500
_MAX_COLS = 20
_XLSX_LARGE_FILE_BYTES = 10 * 1024 * 1024
_CELL_TEXT_MAX = 80


def _trunc(s: str) -> str:
    s = str(s).replace("\n", " ").strip()
    return s[:_CELL_TEXT_MAX] + "…" if len(s) > _CELL_TEXT_MAX else s


def _build_merged_ranges(ws) -> tuple[dict[tuple[int, int], tuple[int, int]], set[tuple[int, int]]]:
    """Return (masters, slaves) where:
    masters: {(r0,c0): (row_span, col_span)} — 0-indexed master positions
    slaves: set of (r,c) — 0-indexed slave positions to skip
    """
    masters: dict[tuple[int, int], tuple[int, int]] = {}
    slaves: set[tuple[int, int]] = set()
    try:
        for rng in ws.merged_cells.ranges:
            mr = rng.min_row - 1
            mc = rng.min_col - 1
            row_span = rng.max_row - rng.min_row + 1
            col_span = rng.max_col - rng.min_col + 1
            masters[(mr, mc)] = (row_span, col_span)
            for row in range(rng.min_row, rng.max_row + 1):
                for col in range(rng.min_col, rng.max_col + 1):
                    if row != rng.min_row or col != rng.min_col:
                        slaves.add((row - 1, col - 1))
    except Exception:
        pass
    return masters, slaves


_KV_HEADER_NAMES = frozenset({
    "name", "value", "key", "field", "property", "attribute",
    "id", "score", "count", "total", "amount", "date", "type",
    "category", "description", "label", "item", "col1", "col2",
})


def _is_kv_region(cells: list, row_count: int, col_count: int) -> bool:
    """Return True if this region looks like a 2-column KV layout, not a data table.

    Criteria:
    - Exactly 2 columns populated
    - Short region (<= 20 rows) — data tables tend to be taller
    - First column: all short text (labels, <= 6 words)
    - Second column: all short values (<= 80 chars)
    - At least 3 rows
    - No repeated first-column values (otherwise it's a data table)
    - First row must not look like a column-header pair (e.g., "Name"/"Value",
      "ID"/"Score" — conventional header names are rejected)
    """
    if col_count != 2:
        return False
    if row_count < 3:
        return False
    # KV layouts are compact; large tables are data tables
    if row_count > 20:
        return False

    # Check if the first row looks like a column-header pair
    first_row: dict[int, str] = {}
    for cell in cells:
        if cell.row == 0:
            first_row[cell.column] = cell.text.strip().lower()

    first_col0 = first_row.get(0, "")
    first_col1 = first_row.get(1, "")
    # If either of the first-row values are generic column header names, it's a data table
    if first_col0 in _KV_HEADER_NAMES or first_col1 in _KV_HEADER_NAMES:
        return False

    col0_values: dict[str, bool] = {}
    for cell in cells:
        if cell.column == 0:
            text = cell.text.strip()
            if not text:
                continue
            # Long labels are likely column headers or prose, not KV labels
            if len(text.split()) > 6:
                return False
            # Repeated first-column values → data table
            if text in col0_values:
                return False
            col0_values[text] = True
        elif cell.column == 1:
            if len(cell.text) > 80:
                return False

    # Must have at least 3 matching key-value pairs
    return len(col0_values) >= 3


def _xlsx_cells_to_kv_group(cells: list, sheet_name: str) -> KeyValueGroup | None:
    """Convert a two-column cell list to a KeyValueGroup."""
    from ...models.key_value import KeyValueEntry, KeyValueGroup, KeyValueGroupType

    # Build row -> (col0_text, col1_text) map
    row_map: dict[int, dict[int, str]] = {}
    for cell in cells:
        row_map.setdefault(cell.row, {})[cell.column] = cell.text.strip()

    entries = []
    for r in sorted(row_map.keys()):
        row = row_map[r]
        key = row.get(0, "").strip()
        value = row.get(1, "").strip()
        if key:
            entries.append(KeyValueEntry(
                key=key,
                value=value,
                confidence="extracted",
            ))

    if len(entries) < 3:
        return None

    return KeyValueGroup(
        entries=entries,
        title=sheet_name,
        group_type=KeyValueGroupType.METADATA,
        extraction_method="xlsx.two_column_region",
        confidence="extracted",
    )


# ── XLSX ──────────────────────────────────────────────────────────────────────

class XlsxParser(ParserPlugin):
    name = "xlsx_parser"
    supported_types = ["xlsx", "xlsm", "xltx", "xltm"]

    def execute(self, ctx: CompilationContext) -> CompilationContext:
        import openpyxl

        path = Path(ctx.source)
        file_size = path.stat().st_size
        read_only = file_size > _XLSX_LARGE_FILE_BYTES

        try:
            # Primary load: cached values for display text
            wb = openpyxl.load_workbook(str(path), read_only=read_only, data_only=True)
        except Exception as e:
            ctx.error("XLSX_PARSE_ERROR", str(e))
            return ctx

        # Secondary load for formulas (non-read-only small files only)
        wb_f = None
        if not read_only:
            try:
                wb_f = openpyxl.load_workbook(str(path), read_only=False, data_only=False)
            except Exception:
                wb_f = None

        blocks: list[Block] = []
        idx = 0

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            ws_f = wb_f[sheet_name] if wb_f is not None else None

            blocks.append(Block(type=BlockType.HEADING, content=sheet_name, level=2, index=idx))
            idx += 1

            masters, slaves = _build_merged_ranges(ws) if not read_only else ({}, set())

            cells: list[TableCell] = []
            col_count = 0
            rows_added = 0

            for r_idx, row in enumerate(ws.iter_rows(max_row=_MAX_ROWS_PER_SHEET + 1)):
                for c_idx, cell in enumerate(row):
                    if c_idx >= _MAX_COLS:
                        break
                    if (r_idx, c_idx) in slaves:
                        continue

                    row_span, col_span = masters.get((r_idx, c_idx), (1, 1))

                    # Get cached display value
                    raw_val = cell.value
                    formula = None
                    data_type = None
                    number_format = None

                    if ws_f is not None:
                        try:
                            fc = ws_f.cell(r_idx + 1, c_idx + 1)
                            data_type = fc.data_type
                            if fc.data_type == 'f':
                                formula = fc.value
                            nf = fc.number_format
                            if nf and nf != 'General':
                                number_format = nf
                        except Exception:
                            pass

                    if raw_val is not None:
                        display = _trunc(str(raw_val))
                    elif formula:
                        display = formula
                    else:
                        display = ""

                    col_count = max(col_count, c_idx + col_span)
                    cells.append(TableCell(
                        text=display,
                        row=r_idx,
                        column=c_idx,
                        row_span=row_span,
                        column_span=col_span,
                        raw_value=raw_val,
                        formula=formula,
                        data_type=data_type or (cell.data_type if hasattr(cell, 'data_type') else None),
                        number_format=number_format,
                    ))
                rows_added = r_idx + 1

            if not cells:
                continue

            # Count populated rows (rows with at least one non-empty cell)
            populated_rows: set[int] = {c.row for c in cells if c.text.strip()}
            populated_row_count = len(populated_rows)

            # Check for two-column key-value layout
            if _is_kv_region(cells, populated_row_count, col_count):
                kv_group = _xlsx_cells_to_kv_group(cells, sheet_name)
                if kv_group is not None:
                    kv_block = Block.from_key_value_group(
                        kv_group,
                        page=None,
                        index=idx,
                        confidence=ExtractionConfidence.EXTRACTED,
                    )
                    blocks.append(kv_block)
                    idx += 1
                    continue  # skip the normal table-block emission for this sheet

            # Omitted row count from worksheet metadata
            actual_max = getattr(ws, 'max_row', None) or rows_added
            omitted = max(0, actual_max - rows_added)

            table = TableData(
                row_count=rows_added,
                column_count=col_count or 1,
                cells=cells,
                header_rows=[0] if rows_added > 0 else [],
                header_detection="assumed_first_row",
                span_detection="native" if not read_only else "unsupported",
                extraction_method=ExtractionMethod.XLSX_NATIVE,
                sheet=sheet_name,
                metadata={
                    "omitted_rows": omitted,
                    "read_only_mode": read_only,
                    **({"formula_available": wb_f is not None} if not read_only else {}),
                },
            )
            blocks.append(Block.from_table(table, index=idx))
            idx += 1

            if omitted > 0:
                blocks.append(Block(
                    type=BlockType.PARAGRAPH,
                    content=f"*({omitted} additional rows omitted)*",
                    index=idx,
                ))
                idx += 1

        if wb_f is not None:
            wb_f.close()
        wb.close()

        ctx.document = Document(
            source=str(path),
            file_type=Path(path).suffix.lstrip(".").lower(),
            title=path.stem,
            pages=len(wb.sheetnames),
            blocks=blocks,
            metadata={"sheets": list(wb.sheetnames)},
        ).compute_id()
        return ctx


# ── XLS (legacy) ──────────────────────────────────────────────────────────────

class XlsParser(ParserPlugin):
    name = "xls_parser"
    supported_types = ["xls"]

    def execute(self, ctx: CompilationContext) -> CompilationContext:
        import xlrd

        path = Path(ctx.source)
        try:
            wb = xlrd.open_workbook(str(path))
        except Exception as e:
            ctx.error("XLS_PARSE_ERROR", str(e))
            return ctx

        blocks: list[Block] = []
        idx = 0

        for sheet in wb.sheets():
            blocks.append(Block(type=BlockType.HEADING, content=sheet.name, level=2, index=idx))
            idx += 1
            if sheet.nrows == 0:
                continue

            ncols = min(sheet.ncols, _MAX_COLS)
            nrows = min(sheet.nrows, _MAX_ROWS_PER_SHEET + 1)

            cells: list[TableCell] = []
            for r_idx in range(nrows):
                for c_idx in range(ncols):
                    val = sheet.cell_value(r_idx, c_idx)
                    cells.append(TableCell(
                        text=_trunc(str(val)) if val is not None and val != '' else "",
                        row=r_idx,
                        column=c_idx,
                        raw_value=val,
                    ))

            omitted = max(0, sheet.nrows - nrows)
            table = TableData(
                row_count=nrows,
                column_count=ncols,
                cells=cells,
                header_rows=[0] if nrows > 0 else [],
                header_detection="assumed_first_row",
                span_detection="unsupported",
                extraction_method=ExtractionMethod.XLS_NATIVE,
                sheet=sheet.name,
                metadata={"omitted_rows": omitted},
            )
            blocks.append(Block.from_table(table, index=idx))
            idx += 1

            if omitted > 0:
                blocks.append(Block(
                    type=BlockType.PARAGRAPH,
                    content=f"*({omitted} additional rows omitted)*",
                    index=idx,
                ))
                idx += 1

        ctx.document = Document(
            source=str(path),
            file_type="xls",
            title=path.stem,
            pages=wb.nsheets,
            blocks=blocks,
            metadata={"sheets": wb.sheet_names()},
        ).compute_id()
        return ctx


# ── CSV ───────────────────────────────────────────────────────────────────────

class CsvParser(ParserPlugin):
    name = "csv_parser"
    supported_types = ["csv", "tsv"]

    def execute(self, ctx: CompilationContext) -> CompilationContext:
        import chardet

        path = Path(ctx.source)
        raw = path.read_bytes()
        enc = chardet.detect(raw).get("encoding") or "utf-8"

        ext = path.suffix.lstrip(".").lower()
        delimiter = "\t" if ext == "tsv" else ","

        try:
            text = raw.decode(enc, errors="replace")
            reader = _csv.reader(io.StringIO(text), delimiter=delimiter)
            all_rows = [row for row in reader if any(c.strip() for c in row)]
        except Exception as e:
            ctx.error("CSV_PARSE_ERROR", str(e))
            return ctx

        if not all_rows:
            ctx.document = Document(source=str(path), file_type=ext, title=path.stem, blocks=[]).compute_id()
            return ctx

        nrows = min(len(all_rows), _MAX_ROWS_PER_SHEET + 1)
        ncols = min(max(len(r) for r in all_rows[:nrows]), _MAX_COLS)

        cells: list[TableCell] = []
        for r_idx, row in enumerate(all_rows[:nrows]):
            for c_idx in range(ncols):
                val = row[c_idx] if c_idx < len(row) else ""
                cells.append(TableCell(
                    text=_trunc(val),
                    row=r_idx,
                    column=c_idx,
                ))

        omitted = len(all_rows) - nrows
        method = ExtractionMethod.TSV_NATIVE if ext == "tsv" else ExtractionMethod.CSV_NATIVE

        table = TableData(
            row_count=nrows,
            column_count=ncols,
            cells=cells,
            header_rows=[0] if nrows > 0 else [],
            header_detection="assumed_first_row",
            span_detection="unsupported",
            extraction_method=method,
            metadata={"omitted_rows": omitted},
        )

        blocks: list[Block] = [
            Block(type=BlockType.METADATA,
                  content=f"File: {path.name} | Columns: {ncols} | Rows: {len(all_rows)-1}",
                  index=0),
            Block.from_table(table, index=1),
        ]
        if omitted > 0:
            blocks.append(Block(type=BlockType.PARAGRAPH,
                                content=f"*({omitted} additional rows omitted from preview)*",
                                index=2))

        ctx.document = Document(
            source=str(path),
            file_type=ext,
            title=path.stem,
            pages=1,
            blocks=blocks,
            metadata={"columns": ncols, "rows": len(all_rows) - 1},
        ).compute_id()
        return ctx


register_parser("xlsx", XlsxParser)
register_parser("xlsm", XlsxParser)
register_parser("xls",  XlsParser)
register_parser("csv",  CsvParser)
register_parser("tsv",  CsvParser)
