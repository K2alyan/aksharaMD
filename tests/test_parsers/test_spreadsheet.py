from __future__ import annotations

from pathlib import Path

import openpyxl

from aksharamd.context import CompilationContext
from aksharamd.models.block import BlockType
from aksharamd.plugins.parsers.spreadsheet import CsvParser, XlsxParser


def _parse_xlsx(path: Path, tmp_path: Path) -> CompilationContext:
    ctx = CompilationContext(source=str(path), output_dir=str(tmp_path / "out"))
    return XlsxParser().execute(ctx)


def _parse_csv(path: Path, tmp_path: Path) -> CompilationContext:
    ctx = CompilationContext(source=str(path), output_dir=str(tmp_path / "out"))
    return CsvParser().execute(ctx)


def _make_xlsx(tmp_path: Path, sheets: dict[str, list[list]]) -> Path:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for sheet_name, rows in sheets.items():
        ws = wb.create_sheet(sheet_name)
        for row in rows:
            ws.append(row)
    path = tmp_path / "test.xlsx"
    wb.save(str(path))
    return path


def test_xlsx_basic(tmp_path):
    path = _make_xlsx(tmp_path, {
        "Sheet1": [["Name", "Score"], ["Alice", 95], ["Bob", 87]]
    })
    ctx = _parse_xlsx(path, tmp_path)
    tables = [b for b in ctx.document.blocks if b.type == BlockType.TABLE]
    assert len(tables) == 1
    assert "Name" in tables[0].content
    assert "Alice" in tables[0].content


def test_xlsx_multiple_sheets(tmp_path):
    path = _make_xlsx(tmp_path, {
        "Q1": [["Month", "Revenue"], ["Jan", 1000]],
        "Q2": [["Month", "Revenue"], ["Apr", 1200]],
    })
    ctx = _parse_xlsx(path, tmp_path)
    headings = [b for b in ctx.document.blocks if b.type == BlockType.HEADING]
    assert any("Q1" in h.content for h in headings)
    assert any("Q2" in h.content for h in headings)
    tables = [b for b in ctx.document.blocks if b.type == BlockType.TABLE]
    assert len(tables) == 2


def test_xlsx_merged_cells(tmp_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Merged"
    ws["A1"] = "Region"
    ws["B1"] = "Sales"
    ws["A2"] = "North"
    ws["B2"] = 100
    ws["A3"] = None  # slave cell of merged A2:A3
    ws["B3"] = 150
    ws.merge_cells("A2:A3")
    path = tmp_path / "merged.xlsx"
    wb.save(str(path))

    ctx = _parse_xlsx(path, tmp_path)
    tables = [b for b in ctx.document.blocks if b.type == BlockType.TABLE]
    assert len(tables) == 1
    # Both rows should show "North" (slave cell expanded)
    assert tables[0].content.count("North") == 2


def test_xlsx_row_cap(tmp_path):
    # Create more than 500 rows
    rows = [["ID", "Value"]] + [[str(i), str(i * 10)] for i in range(600)]
    path = _make_xlsx(tmp_path, {"Big": rows})
    ctx = _parse_xlsx(path, tmp_path)
    tables = [b for b in ctx.document.blocks if b.type == BlockType.TABLE]
    assert len(tables) == 1
    # Should mention omitted rows
    full_content = "\n".join(b.content for b in ctx.document.blocks)
    assert "omitted" in full_content


def test_xlsx_empty_sheet(tmp_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Empty"
    path = tmp_path / "empty.xlsx"
    wb.save(str(path))
    ctx = _parse_xlsx(path, tmp_path)
    assert ctx.document is not None
    # Should have a heading block for the sheet, no table
    headings = [b for b in ctx.document.blocks if b.type == BlockType.HEADING]
    assert any("Empty" in h.content for h in headings)


def test_csv_basic(tmp_path):
    p = tmp_path / "data.csv"
    p.write_text("Name,Age,City\nAlice,30,London\nBob,25,Paris\n", encoding="utf-8")
    ctx = _parse_csv(p, tmp_path)
    tables = [b for b in ctx.document.blocks if b.type == BlockType.TABLE]
    assert len(tables) == 1
    assert "Name" in tables[0].content
    assert "Alice" in tables[0].content


def test_tsv_basic(tmp_path):
    p = tmp_path / "data.tsv"
    p.write_text("Col1\tCol2\nA\tB\nC\tD\n", encoding="utf-8")
    ctx = CompilationContext(source=str(p), output_dir=str(tmp_path / "out"))
    ctx = CsvParser().execute(ctx)
    tables = [b for b in ctx.document.blocks if b.type == BlockType.TABLE]
    assert len(tables) == 1
    assert "Col1" in tables[0].content


def test_csv_row_cap(tmp_path):
    rows = ["id,val"] + [f"{i},{i*2}" for i in range(600)]
    p = tmp_path / "big.csv"
    p.write_text("\n".join(rows), encoding="utf-8")
    ctx = _parse_csv(p, tmp_path)
    full_content = "\n".join(b.content for b in ctx.document.blocks)
    assert "omitted" in full_content
