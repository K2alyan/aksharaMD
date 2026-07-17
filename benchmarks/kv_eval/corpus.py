"""Development corpus for KV detection evaluation.

Programmatic synthetic fixtures — no real documents needed.
All cases are pure Python, no external file dependencies.
"""
from __future__ import annotations

import os
import tempfile
from dataclasses import dataclass, field

from benchmarks.kv_eval.ground_truth import (
    GroundTruthEntry,
    KeyValueGroundTruth,
)
from aksharamd.models.key_value import KeyValueGroupType, KeyValueValueType


@dataclass
class TextCase:
    case_id: str
    text: str
    ground_truth: KeyValueGroundTruth


@dataclass
class HtmlCase:
    case_id: str
    html: str
    ground_truth: KeyValueGroundTruth


@dataclass
class XlsxCase:
    case_id: str
    xlsx_path: str
    ground_truth: KeyValueGroundTruth
    _wb: object = None  # hold reference to prevent GC until path is written


# ── True Positive inline cases ──────────────────────────────────────────────

def _tp_inline_cases() -> list[TextCase]:
    cases = []

    # 1. Contact: email + phone
    cases.append(TextCase(
        case_id="contact_card_01",
        text="Email: alice@example.com\nPhone: 555-1234",
        ground_truth=KeyValueGroundTruth(
            case_id="contact_card_01",
            document_id="contact_card_01",
            source_format="text",
            detection_path="heuristic_inline",
            is_key_value_group=True,
            group_type=KeyValueGroupType.CONTACT,
            records=[[
                GroundTruthEntry(key="Email", value="alice@example.com",
                                 expected_value_type=KeyValueValueType.EMAIL),
                GroundTruthEntry(key="Phone", value="555-1234",
                                 expected_value_type=KeyValueValueType.PHONE),
            ]],
        ),
    ))

    # 2. Contact: company name + address
    cases.append(TextCase(
        case_id="contact_card_02",
        text="Company: Acme Corp\nAddress: 123 Main St",
        ground_truth=KeyValueGroundTruth(
            case_id="contact_card_02",
            document_id="contact_card_02",
            source_format="text",
            detection_path="heuristic_inline",
            is_key_value_group=True,
            group_type=KeyValueGroupType.CONTACT,
        ),
    ))

    # 3. Event: date + time + location + organizer
    cases.append(TextCase(
        case_id="event_card_01",
        text="Date: 15/06/2024\nTime: 6:00 PM\nLocation: Town Hall\nOrganizer: City Council",
        ground_truth=KeyValueGroundTruth(
            case_id="event_card_01",
            document_id="event_card_01",
            source_format="text",
            detection_path="heuristic_inline",
            is_key_value_group=True,
            group_type=KeyValueGroupType.EVENT,
        ),
    ))

    # 4. Metadata: title + author + version + status
    cases.append(TextCase(
        case_id="doc_metadata_01",
        text="Title: Project Charter\nAuthor: Jane Doe\nVersion: 2.1\nStatus: Draft",
        ground_truth=KeyValueGroundTruth(
            case_id="doc_metadata_01",
            document_id="doc_metadata_01",
            source_format="text",
            detection_path="heuristic_inline",
            is_key_value_group=True,
            group_type=KeyValueGroupType.METADATA,
        ),
    ))

    # 5. Specification: voltage + current + power
    cases.append(TextCase(
        case_id="spec_sheet_01",
        text="Voltage: 12V\nCurrent: 2A\nPower: 24W",
        ground_truth=KeyValueGroundTruth(
            case_id="spec_sheet_01",
            document_id="spec_sheet_01",
            source_format="text",
            detection_path="heuristic_inline",
            is_key_value_group=True,
            group_type=KeyValueGroupType.SPECIFICATION,
        ),
    ))

    # 6. Form: name + address + date signed
    cases.append(TextCase(
        case_id="form_01",
        text="Name: John Smith\nAddress: 456 Oak Ave\nDate Signed: 01/03/2024",
        ground_truth=KeyValueGroundTruth(
            case_id="form_01",
            document_id="form_01",
            source_format="text",
            detection_path="heuristic_inline",
            is_key_value_group=True,
            group_type=KeyValueGroupType.FORM,
        ),
    ))

    # 7. Office hours: Mon-Fri weekday
    cases.append(TextCase(
        case_id="office_hours_01",
        text="Monday: 9:00 AM - 5:00 PM\nTuesday: 9:00 AM - 5:00 PM\nWednesday: 9:00 AM - 5:00 PM",
        ground_truth=KeyValueGroundTruth(
            case_id="office_hours_01",
            document_id="office_hours_01",
            source_format="text",
            detection_path="heuristic_inline",
            is_key_value_group=True,
            group_type=KeyValueGroupType.SCHEDULE,
        ),
    ))

    # 8. Product: model + manufacturer + warranty
    cases.append(TextCase(
        case_id="product_01",
        text="Model: XR-500\nManufacturer: TechCo\nWarranty: 2 years",
        ground_truth=KeyValueGroundTruth(
            case_id="product_01",
            document_id="product_01",
            source_format="text",
            detection_path="heuristic_inline",
            is_key_value_group=True,
            group_type=KeyValueGroupType.SPECIFICATION,
        ),
    ))

    # 9. Government: department + agency + contact
    cases.append(TextCase(
        case_id="government_01",
        text="Department: Health Services\nAgency: State Government\nContact: 1800-555-0001",
        ground_truth=KeyValueGroundTruth(
            case_id="government_01",
            document_id="government_01",
            source_format="text",
            detection_path="heuristic_inline",
            is_key_value_group=True,
            group_type=KeyValueGroupType.CONTACT,
        ),
    ))

    # 10. HR: employee + department + title + start date
    cases.append(TextCase(
        case_id="hr_record_01",
        text="Employee: Bob Jones\nDepartment: Engineering\nTitle: Senior Engineer\nStart Date: 2022-01-15",
        ground_truth=KeyValueGroundTruth(
            case_id="hr_record_01",
            document_id="hr_record_01",
            source_format="text",
            detection_path="heuristic_inline",
            is_key_value_group=True,
            group_type=KeyValueGroupType.METADATA,
        ),
    ))

    # 11. Church schedule Saturday service
    cases.append(TextCase(
        case_id="church_schedule_sat",
        text="Day: Saturday\nTime: 7:00 PM\nLocation: St. Mary's Church",
        ground_truth=KeyValueGroundTruth(
            case_id="church_schedule_sat",
            document_id="church_schedule_sat",
            source_format="text",
            detection_path="heuristic_inline",
            is_key_value_group=True,
            group_type=KeyValueGroupType.SCHEDULE,
        ),
    ))

    # 12. Church schedule Sunday service
    cases.append(TextCase(
        case_id="church_schedule_sun",
        text="Day: Sunday\nTime: 9:00 AM\nLocation: St. Mary's Church",
        ground_truth=KeyValueGroundTruth(
            case_id="church_schedule_sun",
            document_id="church_schedule_sun",
            source_format="text",
            detection_path="heuristic_inline",
            is_key_value_group=True,
            group_type=KeyValueGroupType.SCHEDULE,
        ),
    ))

    # 13. Software release metadata
    cases.append(TextCase(
        case_id="release_metadata_01",
        text="Package: aksharamd\nVersion: 0.3.6\nRelease Date: 2024-07-01\nLicense: MIT",
        ground_truth=KeyValueGroundTruth(
            case_id="release_metadata_01",
            document_id="release_metadata_01",
            source_format="text",
            detection_path="heuristic_inline",
            is_key_value_group=True,
            group_type=KeyValueGroupType.METADATA,
        ),
    ))

    # 14. Network device specification
    cases.append(TextCase(
        case_id="network_spec_01",
        text="Hostname: router01\nIP: 192.168.1.1\nSubnet: 255.255.255.0\nGateway: 192.168.1.254",
        ground_truth=KeyValueGroundTruth(
            case_id="network_spec_01",
            document_id="network_spec_01",
            source_format="text",
            detection_path="heuristic_inline",
            is_key_value_group=True,
            group_type=KeyValueGroupType.SPECIFICATION,
        ),
    ))

    # 15. Project metadata with owner
    cases.append(TextCase(
        case_id="project_metadata_01",
        text="Project: AksharaMD\nOwner: Kalyan\nPriority: High\nDeadline: 2024-09-01",
        ground_truth=KeyValueGroundTruth(
            case_id="project_metadata_01",
            document_id="project_metadata_01",
            source_format="text",
            detection_path="heuristic_inline",
            is_key_value_group=True,
            group_type=KeyValueGroupType.METADATA,
        ),
    ))

    # 16. Medical record
    cases.append(TextCase(
        case_id="medical_record_01",
        text="Patient: Mary Johnson\nDOB: 12/05/1985\nBlood Type: O+\nAllergies: None",
        ground_truth=KeyValueGroundTruth(
            case_id="medical_record_01",
            document_id="medical_record_01",
            source_format="text",
            detection_path="heuristic_inline",
            is_key_value_group=True,
            group_type=KeyValueGroupType.FORM,
        ),
    ))

    # 17. Conference room booking
    cases.append(TextCase(
        case_id="room_booking_01",
        text="Room: Conference A\nDate: 20/07/2024\nTime: 2:00 PM\nBooked By: Sarah Lee",
        ground_truth=KeyValueGroundTruth(
            case_id="room_booking_01",
            document_id="room_booking_01",
            source_format="text",
            detection_path="heuristic_inline",
            is_key_value_group=True,
            group_type=KeyValueGroupType.EVENT,
        ),
    ))

    # 18. Vehicle registration
    cases.append(TextCase(
        case_id="vehicle_reg_01",
        text="Make: Toyota\nModel: Corolla\nYear: 2020\nPlate: ABC-1234",
        ground_truth=KeyValueGroundTruth(
            case_id="vehicle_reg_01",
            document_id="vehicle_reg_01",
            source_format="text",
            detection_path="heuristic_inline",
            is_key_value_group=True,
            group_type=KeyValueGroupType.SPECIFICATION,
        ),
    ))

    # 19. Library catalog entry
    cases.append(TextCase(
        case_id="library_catalog_01",
        text="ISBN: 978-0-316-76948-0\nTitle: The Catcher in the Rye\nAuthor: J.D. Salinger\nYear: 1951",
        ground_truth=KeyValueGroundTruth(
            case_id="library_catalog_01",
            document_id="library_catalog_01",
            source_format="text",
            detection_path="heuristic_inline",
            is_key_value_group=True,
            group_type=KeyValueGroupType.METADATA,
        ),
    ))

    # 20. Weather report
    cases.append(TextCase(
        case_id="weather_01",
        text="Location: Sydney\nTemp: 24C\nHumidity: 65%\nWind: 15 km/h",
        ground_truth=KeyValueGroundTruth(
            case_id="weather_01",
            document_id="weather_01",
            source_format="text",
            detection_path="heuristic_inline",
            is_key_value_group=True,
            group_type=KeyValueGroupType.SPECIFICATION,
        ),
    ))

    # 21. API endpoint metadata
    cases.append(TextCase(
        case_id="api_endpoint_01",
        text="Method: GET\nPath: /api/v1/users\nAuth: Bearer token\nRate Limit: 100/min",
        ground_truth=KeyValueGroundTruth(
            case_id="api_endpoint_01",
            document_id="api_endpoint_01",
            source_format="text",
            detection_path="heuristic_inline",
            is_key_value_group=True,
            group_type=KeyValueGroupType.SPECIFICATION,
        ),
    ))

    # 22. Property listing
    cases.append(TextCase(
        case_id="property_01",
        text="Address: 789 Pine St\nPrice: $450,000\nBedrooms: 3\nBathrooms: 2",
        ground_truth=KeyValueGroundTruth(
            case_id="property_01",
            document_id="property_01",
            source_format="text",
            detection_path="heuristic_inline",
            is_key_value_group=True,
            group_type=KeyValueGroupType.UNKNOWN,
        ),
    ))

    # 23. Invoice header
    cases.append(TextCase(
        case_id="invoice_header_01",
        text="Invoice No: INV-2024-001\nDate: 01/07/2024\nClient: ABC Ltd\nAmount: $1,250",
        ground_truth=KeyValueGroundTruth(
            case_id="invoice_header_01",
            document_id="invoice_header_01",
            source_format="text",
            detection_path="heuristic_inline",
            is_key_value_group=True,
            group_type=KeyValueGroupType.FORM,
        ),
    ))

    # 24. Research paper metadata
    cases.append(TextCase(
        case_id="paper_metadata_01",
        text="Title: Deep Learning Survey\nAuthors: Smith, Jones\nJournal: Nature\nYear: 2023",
        ground_truth=KeyValueGroundTruth(
            case_id="paper_metadata_01",
            document_id="paper_metadata_01",
            source_format="text",
            detection_path="heuristic_inline",
            is_key_value_group=True,
            group_type=KeyValueGroupType.METADATA,
        ),
    ))

    # 25. Server configuration
    cases.append(TextCase(
        case_id="server_config_01",
        text="Host: db.example.com\nPort: 5432\nDatabase: production\nSSL: enabled",
        ground_truth=KeyValueGroundTruth(
            case_id="server_config_01",
            document_id="server_config_01",
            source_format="text",
            detection_path="heuristic_inline",
            is_key_value_group=True,
            group_type=KeyValueGroupType.SPECIFICATION,
        ),
    ))

    # 26. Training schedule
    cases.append(TextCase(
        case_id="training_schedule_01",
        text="Monday: 6:00 AM\nWednesday: 6:00 AM\nFriday: 7:00 AM",
        ground_truth=KeyValueGroundTruth(
            case_id="training_schedule_01",
            document_id="training_schedule_01",
            source_format="text",
            detection_path="heuristic_inline",
            is_key_value_group=True,
            group_type=KeyValueGroupType.SCHEDULE,
        ),
    ))

    # 27. Flight details
    cases.append(TextCase(
        case_id="flight_details_01",
        text="Flight: QF001\nFrom: Sydney\nTo: London\nDeparture: 10:30 AM",
        ground_truth=KeyValueGroundTruth(
            case_id="flight_details_01",
            document_id="flight_details_01",
            source_format="text",
            detection_path="heuristic_inline",
            is_key_value_group=True,
            group_type=KeyValueGroupType.EVENT,
        ),
    ))

    # 28. Nutritional info
    cases.append(TextCase(
        case_id="nutrition_01",
        text="Calories: 250\nProtein: 12g\nCarbs: 30g\nFat: 8g",
        ground_truth=KeyValueGroundTruth(
            case_id="nutrition_01",
            document_id="nutrition_01",
            source_format="text",
            detection_path="heuristic_inline",
            is_key_value_group=True,
            group_type=KeyValueGroupType.SPECIFICATION,
        ),
    ))

    # 29. Software license
    cases.append(TextCase(
        case_id="license_01",
        text="License: MIT\nCopyright: 2024 Acme Corp\nExpiry: 2025-12-31",
        ground_truth=KeyValueGroundTruth(
            case_id="license_01",
            document_id="license_01",
            source_format="text",
            detection_path="heuristic_inline",
            is_key_value_group=True,
            group_type=KeyValueGroupType.METADATA,
        ),
    ))

    # 30. Customer support ticket
    cases.append(TextCase(
        case_id="support_ticket_01",
        text="Ticket: SUP-12345\nPriority: High\nAssignee: Tech Team\nStatus: Open",
        ground_truth=KeyValueGroundTruth(
            case_id="support_ticket_01",
            document_id="support_ticket_01",
            source_format="text",
            detection_path="heuristic_inline",
            is_key_value_group=True,
            group_type=KeyValueGroupType.METADATA,
        ),
    ))

    # ── kv_promoter/v2 expanded positive cases ────────────────────────────
    # These target: multilingual contact cards, typed spec sheets, invoice
    # headers with dates/currencies, weekly schedules with 3+ days,
    # medical test results with typed values, product specs with typed
    # units. They exercise Rule A (>=2 strongly typed values) and Rule B
    # (>=3 schema fields).

    # 31. French contact card — Rule B (contact schema)
    cases.append(TextCase(
        case_id="fr_contact_01",
        text="Name: Jean Dupont\nEmail: jean@exemple.fr\nPhone: +33 1 23 45 67 89",
        ground_truth=KeyValueGroundTruth(
            case_id="fr_contact_01",
            document_id="fr_contact_01",
            source_format="text",
            detection_path="heuristic_inline",
            is_key_value_group=True,
            group_type=KeyValueGroupType.CONTACT,
        ),
    ))

    # 32. Spanish contact card
    cases.append(TextCase(
        case_id="es_contact_01",
        text="Name: Maria Garcia\nEmail: maria@ejemplo.es\nPhone: +34 91 123 45 67",
        ground_truth=KeyValueGroundTruth(
            case_id="es_contact_01",
            document_id="es_contact_01",
            source_format="text",
            detection_path="heuristic_inline",
            is_key_value_group=True,
            group_type=KeyValueGroupType.CONTACT,
        ),
    ))

    # 33. Weekly schedule with 3 days
    cases.append(TextCase(
        case_id="weekly_schedule_01",
        text="Monday: 9:00 AM\nTuesday: 10:00 AM\nWednesday: 11:00 AM",
        ground_truth=KeyValueGroundTruth(
            case_id="weekly_schedule_01",
            document_id="weekly_schedule_01",
            source_format="text",
            detection_path="heuristic_inline",
            is_key_value_group=True,
            group_type=KeyValueGroupType.SCHEDULE,
        ),
    ))

    # 34. Weekly schedule with 5 days — full week schedule
    cases.append(TextCase(
        case_id="weekly_schedule_02",
        text="Monday: 9:00 AM\nTuesday: 9:00 AM\nWednesday: 9:00 AM\nThursday: 9:00 AM\nFriday: 9:00 AM",
        ground_truth=KeyValueGroundTruth(
            case_id="weekly_schedule_02",
            document_id="weekly_schedule_02",
            source_format="text",
            detection_path="heuristic_inline",
            is_key_value_group=True,
            group_type=KeyValueGroupType.SCHEDULE,
        ),
    ))

    # 35. Product spec sheet — typed physical units — Rule A + Rule B
    cases.append(TextCase(
        case_id="product_spec_typed_01",
        text="Voltage: 12V\nCurrent: 2A\nFrequency: 60Hz",
        ground_truth=KeyValueGroundTruth(
            case_id="product_spec_typed_01",
            document_id="product_spec_typed_01",
            source_format="text",
            detection_path="heuristic_inline",
            is_key_value_group=True,
            group_type=KeyValueGroupType.SPECIFICATION,
        ),
    ))

    # 36. Invoice header — dates + currency (Rule A: >=2 strongly typed)
    cases.append(TextCase(
        case_id="invoice_header_v2_01",
        text="Invoice: INV-2024-001\nDate: 15/06/2024\nDue: 15/07/2024\nTotal: $1,250.00",
        ground_truth=KeyValueGroundTruth(
            case_id="invoice_header_v2_01",
            document_id="invoice_header_v2_01",
            source_format="text",
            detection_path="heuristic_inline",
            is_key_value_group=True,
            group_type=KeyValueGroupType.FORM,
        ),
    ))

    # 37. Medical test results — typed values
    cases.append(TextCase(
        case_id="lab_results_v2_01",
        text="Glucose: 5.2\nHemoglobin: 14.5\nDate: 12/06/2024\nTime: 9:00 AM",
        ground_truth=KeyValueGroundTruth(
            case_id="lab_results_v2_01",
            document_id="lab_results_v2_01",
            source_format="text",
            detection_path="heuristic_inline",
            is_key_value_group=True,
            group_type=KeyValueGroupType.METADATA,
        ),
    ))

    # 38. Contact with website — Rule A (>=2 strong: url + phone)
    cases.append(TextCase(
        case_id="contact_web_01",
        text="Website: https://example.com\nPhone: 555-2000\nAddress: 100 Broadway",
        ground_truth=KeyValueGroundTruth(
            case_id="contact_web_01",
            document_id="contact_web_01",
            source_format="text",
            detection_path="heuristic_inline",
            is_key_value_group=True,
            group_type=KeyValueGroupType.CONTACT,
        ),
    ))

    # 39. Booking confirmation — event with typed values
    cases.append(TextCase(
        case_id="booking_confirmation_01",
        text="Booking: BK-1234\nDate: 20/07/2024\nTime: 8:00 PM\nVenue: Grand Hall",
        ground_truth=KeyValueGroundTruth(
            case_id="booking_confirmation_01",
            document_id="booking_confirmation_01",
            source_format="text",
            detection_path="heuristic_inline",
            is_key_value_group=True,
            group_type=KeyValueGroupType.EVENT,
        ),
    ))

    # 40. Package release metadata — Rule B
    cases.append(TextCase(
        case_id="package_release_01",
        text="Package: aksharamd\nVersion: 0.4.0\nAuthor: Kalyan\nLicense: MIT\nPublisher: aksharamd-team",
        ground_truth=KeyValueGroundTruth(
            case_id="package_release_01",
            document_id="package_release_01",
            source_format="text",
            detection_path="heuristic_inline",
            is_key_value_group=True,
            group_type=KeyValueGroupType.METADATA,
        ),
    ))

    # 41. Multilingual DE contact
    cases.append(TextCase(
        case_id="de_contact_01",
        text="Name: Hans Müller\nEmail: hans@beispiel.de\nPhone: +49 30 12345678",
        ground_truth=KeyValueGroundTruth(
            case_id="de_contact_01",
            document_id="de_contact_01",
            source_format="text",
            detection_path="heuristic_inline",
            is_key_value_group=True,
            group_type=KeyValueGroupType.CONTACT,
        ),
    ))

    # 42. Percentages + numbers KPI (strongly typed)
    cases.append(TextCase(
        case_id="kpi_report_01",
        text="Revenue: $1.2M\nGrowth: 15%\nMargin: 22%",
        ground_truth=KeyValueGroundTruth(
            case_id="kpi_report_01",
            document_id="kpi_report_01",
            source_format="text",
            detection_path="heuristic_inline",
            is_key_value_group=True,
            group_type=KeyValueGroupType.UNKNOWN,
        ),
    ))

    # 43. Warranty spec sheet — Rule B
    cases.append(TextCase(
        case_id="warranty_spec_01",
        text="Model: WR-2024\nManufacturer: WarrantyCo\nWarranty: 3 years\nSerial: SN-9988",
        ground_truth=KeyValueGroundTruth(
            case_id="warranty_spec_01",
            document_id="warranty_spec_01",
            source_format="text",
            detection_path="heuristic_inline",
            is_key_value_group=True,
            group_type=KeyValueGroupType.SPECIFICATION,
        ),
    ))

    # 44. Conference talk metadata — Rule B (event)
    cases.append(TextCase(
        case_id="talk_metadata_01",
        text="Event: PyCon\nDate: 15/05/2024\nTime: 2:00 PM\nLocation: Main Hall",
        ground_truth=KeyValueGroundTruth(
            case_id="talk_metadata_01",
            document_id="talk_metadata_01",
            source_format="text",
            detection_path="heuristic_inline",
            is_key_value_group=True,
            group_type=KeyValueGroupType.EVENT,
        ),
    ))

    # 45. Simple contact — Rule A soft rule (1 typed + 2 schema)
    cases.append(TextCase(
        case_id="soft_contact_01",
        text="Name: Alice Chen\nEmail: alice@corp.com\nCompany: Acme",
        ground_truth=KeyValueGroundTruth(
            case_id="soft_contact_01",
            document_id="soft_contact_01",
            source_format="text",
            detection_path="heuristic_inline",
            is_key_value_group=True,
            group_type=KeyValueGroupType.CONTACT,
        ),
    ))

    # 46. Small business hours (schedule 3 fields)
    cases.append(TextCase(
        case_id="business_hours_01",
        text="Opening: 8:00 AM\nClosing: 6:00 PM\nDay: Weekdays",
        ground_truth=KeyValueGroundTruth(
            case_id="business_hours_01",
            document_id="business_hours_01",
            source_format="text",
            detection_path="heuristic_inline",
            is_key_value_group=True,
            group_type=KeyValueGroupType.SCHEDULE,
        ),
    ))

    # 47. Journal article citation — bibliographic KV with schema signals
    # NOTE: this has Author/Year/Journal keys, but Year is a plain 4-digit — the
    # citation exclusion needs 2+ year values, this only has 1.
    cases.append(TextCase(
        case_id="journal_article_01",
        text="Title: Quantum Mechanics\nAuthor: Feynman\nJournal: Physical Review\nYear: 1965",
        ground_truth=KeyValueGroundTruth(
            case_id="journal_article_01",
            document_id="journal_article_01",
            source_format="text",
            detection_path="heuristic_inline",
            is_key_value_group=True,
            group_type=KeyValueGroupType.METADATA,
        ),
    ))

    # 48. Multi-phone company — Rule A (>=2 phones)
    cases.append(TextCase(
        case_id="multi_phone_01",
        text="Sales: +1-555-0100\nSupport: +1-555-0200\nFax: +1-555-0300",
        ground_truth=KeyValueGroundTruth(
            case_id="multi_phone_01",
            document_id="multi_phone_01",
            source_format="text",
            detection_path="heuristic_inline",
            is_key_value_group=True,
            group_type=KeyValueGroupType.CONTACT,
        ),
    ))

    # 49. Server capacity spec — Rule B (spec fields)
    cases.append(TextCase(
        case_id="server_spec_01",
        text="Manufacturer: DellEMC\nModel: R650\nWarranty: 5 years\nCapacity: 24 TB",
        ground_truth=KeyValueGroundTruth(
            case_id="server_spec_01",
            document_id="server_spec_01",
            source_format="text",
            detection_path="heuristic_inline",
            is_key_value_group=True,
            group_type=KeyValueGroupType.SPECIFICATION,
        ),
    ))

    # 50. Registration form (form schema, Rule B)
    cases.append(TextCase(
        case_id="registration_form_01",
        text="Reference: REG-9999\nCustomer: Bob\nAmount: $50\nStatus: Paid",
        ground_truth=KeyValueGroundTruth(
            case_id="registration_form_01",
            document_id="registration_form_01",
            source_format="text",
            detection_path="heuristic_inline",
            is_key_value_group=True,
            group_type=KeyValueGroupType.FORM,
        ),
    ))

    return cases


# ── Negative control cases ────────────────────────────────────────────────────

def _negative_control_cases() -> list[TextCase]:
    cases = []

    # N1. Rhetorical colon: "Note: ..."
    cases.append(TextCase(
        case_id="rhetorical_prose_01",
        text="Note: this explains the process in detail and should be read carefully before proceeding.",
        ground_truth=KeyValueGroundTruth(
            case_id="rhetorical_prose_01",
            document_id="rhetorical_prose_01",
            source_format="text",
            detection_path="negative_control",
            is_key_value_group=False,
            negative_reason="rhetorical_label",
        ),
    ))

    # N2. Rhetorical: "The result was clear: ..."
    cases.append(TextCase(
        case_id="rhetorical_result_01",
        text="Result: the analysis showed significant correlation between the variables measured.",
        ground_truth=KeyValueGroundTruth(
            case_id="rhetorical_result_01",
            document_id="rhetorical_result_01",
            source_format="text",
            detection_path="negative_control",
            is_key_value_group=False,
            negative_reason="rhetorical_label",
        ),
    ))

    # N3. Long value exceeding 80 chars
    cases.append(TextCase(
        case_id="long_value_01",
        text="Description: This is a very long prose sentence that exceeds eighty characters in length and should be rejected.",
        ground_truth=KeyValueGroundTruth(
            case_id="long_value_01",
            document_id="long_value_01",
            source_format="text",
            detection_path="negative_control",
            is_key_value_group=False,
            negative_reason="value_too_long",
        ),
    ))

    # N4. Single entry (non-contact)
    cases.append(TextCase(
        case_id="single_entry_01",
        text="Version: 1.0",
        ground_truth=KeyValueGroundTruth(
            case_id="single_entry_01",
            document_id="single_entry_01",
            source_format="text",
            detection_path="negative_control",
            is_key_value_group=False,
            negative_reason="single_entry_insufficient_evidence",
        ),
    ))

    # N5. Heading followed by paragraph
    cases.append(TextCase(
        case_id="heading_paragraph_01",
        text="Introduction:\nThis section covers the background of the project and its goals.",
        ground_truth=KeyValueGroundTruth(
            case_id="heading_paragraph_01",
            document_id="heading_paragraph_01",
            source_format="text",
            detection_path="negative_control",
            is_key_value_group=False,
            negative_reason="value_too_long",
        ),
    ))

    # N6. Dialogue
    cases.append(TextCase(
        case_id="dialogue_01",
        text="She said: 'Hello.' He replied: 'Goodbye.' They parted ways.",
        ground_truth=KeyValueGroundTruth(
            case_id="dialogue_01",
            document_id="dialogue_01",
            source_format="text",
            detection_path="negative_control",
            is_key_value_group=False,
            negative_reason="prose_not_kv",
        ),
    ))

    # N7. Legal prose with values that exceed the 80-char limit
    cases.append(TextCase(
        case_id="legal_prose_01",
        text="Section 1: All parties agree that the terms and conditions stated herein shall be fully binding and enforceable.\nSection 2: Any disputes arising under this agreement shall be resolved exclusively through binding arbitration proceedings.",
        ground_truth=KeyValueGroundTruth(
            case_id="legal_prose_01",
            document_id="legal_prose_01",
            source_format="text",
            detection_path="negative_control",
            is_key_value_group=False,
            negative_reason="value_too_long",
        ),
    ))

    # N8. Academic rhetorical
    cases.append(TextCase(
        case_id="academic_however_01",
        text="However: the results indicate a correlation between temperature and pressure.",
        ground_truth=KeyValueGroundTruth(
            case_id="academic_however_01",
            document_id="academic_however_01",
            source_format="text",
            detection_path="negative_control",
            is_key_value_group=False,
            negative_reason="rhetorical_label",
        ),
    ))

    # N9. Citation
    cases.append(TextCase(
        case_id="citation_01",
        text="See: Smith et al. 2023 for details on this topic and related experimental findings.",
        ground_truth=KeyValueGroundTruth(
            case_id="citation_01",
            document_id="citation_01",
            source_format="text",
            detection_path="negative_control",
            is_key_value_group=False,
            negative_reason="rhetorical_label",
        ),
    ))

    # N10. Mathematical expression with arrow (label is "where f(x)" — too long / unusual)
    cases.append(TextCase(
        case_id="math_where_01",
        text="The function is defined on the reals.\nwhere: the domain is restricted to positive numbers only",
        ground_truth=KeyValueGroundTruth(
            case_id="math_where_01",
            document_id="math_where_01",
            source_format="text",
            detection_path="negative_control",
            is_key_value_group=False,
            negative_reason="value_too_long",
        ),
    ))

    # N11. Warning rhetorical
    cases.append(TextCase(
        case_id="warning_prose_01",
        text="Warning: do not attempt to modify this file while the service is running as it may cause data loss.",
        ground_truth=KeyValueGroundTruth(
            case_id="warning_prose_01",
            document_id="warning_prose_01",
            source_format="text",
            detection_path="negative_control",
            is_key_value_group=False,
            negative_reason="rhetorical_label",
        ),
    ))

    # N12. Summary rhetorical
    cases.append(TextCase(
        case_id="summary_prose_01",
        text="Summary: this document outlines the key steps required to complete the onboarding process.",
        ground_truth=KeyValueGroundTruth(
            case_id="summary_prose_01",
            document_id="summary_prose_01",
            source_format="text",
            detection_path="negative_control",
            is_key_value_group=False,
            negative_reason="rhetorical_label",
        ),
    ))

    # N13. "The result was:" prose
    cases.append(TextCase(
        case_id="the_result_01",
        text="The result: a comprehensive analysis that required six months of iterative experimentation.",
        ground_truth=KeyValueGroundTruth(
            case_id="the_result_01",
            document_id="the_result_01",
            source_format="text",
            detection_path="negative_control",
            is_key_value_group=False,
            negative_reason="prose_starter",
        ),
    ))

    # N14. Caution rhetorical
    cases.append(TextCase(
        case_id="caution_prose_01",
        text="Caution: high voltage present. Do not open the panel without proper training.",
        ground_truth=KeyValueGroundTruth(
            case_id="caution_prose_01",
            document_id="caution_prose_01",
            source_format="text",
            detection_path="negative_control",
            is_key_value_group=False,
            negative_reason="rhetorical_label",
        ),
    ))

    # N15. Overview rhetorical
    cases.append(TextCase(
        case_id="overview_prose_01",
        text="Overview: this chapter provides a broad introduction to the concepts covered throughout the book.",
        ground_truth=KeyValueGroundTruth(
            case_id="overview_prose_01",
            document_id="overview_prose_01",
            source_format="text",
            detection_path="negative_control",
            is_key_value_group=False,
            negative_reason="rhetorical_label",
        ),
    ))

    # N16. Long prose with colon in middle
    cases.append(TextCase(
        case_id="long_prose_colon_01",
        text="The analysis: demonstrated that the proposed method significantly outperforms all previous baseline approaches.\nConclusion: further study is required to validate these results under real-world conditions.",
        ground_truth=KeyValueGroundTruth(
            case_id="long_prose_colon_01",
            document_id="long_prose_colon_01",
            source_format="text",
            detection_path="negative_control",
            is_key_value_group=False,
            negative_reason="rhetorical_label_and_long_value",
        ),
    ))

    # N17. Purpose rhetorical
    cases.append(TextCase(
        case_id="purpose_prose_01",
        text="Purpose: to enable seamless document parsing across all supported file formats with high accuracy.",
        ground_truth=KeyValueGroundTruth(
            case_id="purpose_prose_01",
            document_id="purpose_prose_01",
            source_format="text",
            detection_path="negative_control",
            is_key_value_group=False,
            negative_reason="rhetorical_label",
        ),
    ))

    # N18. Objective rhetorical
    cases.append(TextCase(
        case_id="objective_prose_01",
        text="Objective: deliver a production-ready key-value extraction pipeline with measurable precision above 90%.",
        ground_truth=KeyValueGroundTruth(
            case_id="objective_prose_01",
            document_id="objective_prose_01",
            source_format="text",
            detection_path="negative_control",
            is_key_value_group=False,
            negative_reason="rhetorical_label",
        ),
    ))

    # N19. Conclusion rhetorical
    cases.append(TextCase(
        case_id="conclusion_prose_01",
        text="Conclusion: the evidence strongly supports the hypothesis, as demonstrated by multiple independent studies.",
        ground_truth=KeyValueGroundTruth(
            case_id="conclusion_prose_01",
            document_id="conclusion_prose_01",
            source_format="text",
            detection_path="negative_control",
            is_key_value_group=False,
            negative_reason="rhetorical_label",
        ),
    ))

    # N20. Background rhetorical
    cases.append(TextCase(
        case_id="background_prose_01",
        text="Background: this project was initiated in response to growing demand for automated document processing.",
        ground_truth=KeyValueGroundTruth(
            case_id="background_prose_01",
            document_id="background_prose_01",
            source_format="text",
            detection_path="negative_control",
            is_key_value_group=False,
            negative_reason="rhetorical_label",
        ),
    ))

    # N21. Context rhetorical
    cases.append(TextCase(
        case_id="context_prose_01",
        text="Context: the organization has been processing documents manually for over a decade.",
        ground_truth=KeyValueGroundTruth(
            case_id="context_prose_01",
            document_id="context_prose_01",
            source_format="text",
            detection_path="negative_control",
            is_key_value_group=False,
            negative_reason="rhetorical_label",
        ),
    ))

    # N22. Example rhetorical
    cases.append(TextCase(
        case_id="example_prose_01",
        text="Example: consider a PDF document with multiple tables, images, and complex multi-column layouts.",
        ground_truth=KeyValueGroundTruth(
            case_id="example_prose_01",
            document_id="example_prose_01",
            source_format="text",
            detection_path="negative_control",
            is_key_value_group=False,
            negative_reason="rhetorical_label",
        ),
    ))

    # N23. Therefore rhetorical
    cases.append(TextCase(
        case_id="therefore_prose_01",
        text="Therefore: we recommend proceeding with option B as it minimizes risk and reduces implementation time.",
        ground_truth=KeyValueGroundTruth(
            case_id="therefore_prose_01",
            document_id="therefore_prose_01",
            source_format="text",
            detection_path="negative_control",
            is_key_value_group=False,
            negative_reason="rhetorical_label",
        ),
    ))

    # N24. Tip rhetorical
    cases.append(TextCase(
        case_id="tip_prose_01",
        text="Tip: always back up your data before running database migrations or schema changes.",
        ground_truth=KeyValueGroundTruth(
            case_id="tip_prose_01",
            document_id="tip_prose_01",
            source_format="text",
            detection_path="negative_control",
            is_key_value_group=False,
            negative_reason="rhetorical_label",
        ),
    ))

    # N25. Important rhetorical
    cases.append(TextCase(
        case_id="important_prose_01",
        text="Important: ensure all dependencies are installed before running the setup script.",
        ground_truth=KeyValueGroundTruth(
            case_id="important_prose_01",
            document_id="important_prose_01",
            source_format="text",
            detection_path="negative_control",
            is_key_value_group=False,
            negative_reason="rhetorical_label",
        ),
    ))

    # N26. Long first value in otherwise short block
    cases.append(TextCase(
        case_id="long_first_value_01",
        text="Description: This feature enables the system to automatically detect and classify key-value patterns.\nStatus: Active",
        ground_truth=KeyValueGroundTruth(
            case_id="long_first_value_01",
            document_id="long_first_value_01",
            source_format="text",
            detection_path="negative_control",
            is_key_value_group=False,
            negative_reason="value_too_long_partial",
            notes="First line has long value; detector filters out that line; only 1 candidate remains",
        ),
    ))

    # N27. Prose starting with "In the"
    cases.append(TextCase(
        case_id="in_the_01",
        text="In the beginning: the system was designed to handle only text documents.\nAt the time: no support for binary formats existed.",
        ground_truth=KeyValueGroundTruth(
            case_id="in_the_01",
            document_id="in_the_01",
            source_format="text",
            detection_path="negative_control",
            is_key_value_group=False,
            negative_reason="prose_starter",
        ),
    ))

    # N28. Short paragraph with no colon
    cases.append(TextCase(
        case_id="no_colon_01",
        text="This document describes the system architecture.\nIt covers deployment, monitoring, and maintenance.",
        ground_truth=KeyValueGroundTruth(
            case_id="no_colon_01",
            document_id="no_colon_01",
            source_format="text",
            detection_path="negative_control",
            is_key_value_group=False,
            negative_reason="no_candidate_pairs",
        ),
    ))

    # N29. Only whitespace / empty
    cases.append(TextCase(
        case_id="empty_text_01",
        text="   \n   \n   ",
        ground_truth=KeyValueGroundTruth(
            case_id="empty_text_01",
            document_id="empty_text_01",
            source_format="text",
            detection_path="negative_control",
            is_key_value_group=False,
            negative_reason="no_candidate_pairs",
        ),
    ))

    # N30. Hybrid: first line is KV, second is long prose (mixed)
    cases.append(TextCase(
        case_id="hybrid_mixed_01",
        text="Status: Active\nDetails: The system is currently running in production with full feature support enabled across all regions.",
        ground_truth=KeyValueGroundTruth(
            case_id="hybrid_mixed_01",
            document_id="hybrid_mixed_01",
            source_format="text",
            detection_path="negative_control",
            is_key_value_group=False,
            negative_reason="value_too_long_partial",
            notes="Second line value is >80 chars and filtered; only 1 candidate passes → single_entry_insufficient_evidence",
        ),
    ))

    return cases


# ── HTML DL cases ─────────────────────────────────────────────────────────────

def _html_positive_cases() -> list[HtmlCase]:
    cases = []

    # H1. Basic 2-pair DL
    cases.append(HtmlCase(
        case_id="html_dl_basic_01",
        html="<dl><dt>Email</dt><dd>alice@example.com</dd><dt>Phone</dt><dd>555-1234</dd></dl>",
        ground_truth=KeyValueGroundTruth(
            case_id="html_dl_basic_01",
            document_id="html_dl_basic_01",
            source_format="html",
            detection_path="native_html_dl",
            is_key_value_group=True,
            group_type=KeyValueGroupType.CONTACT,
        ),
    ))

    # H2. DL with 3+ fields
    cases.append(HtmlCase(
        case_id="html_dl_3fields_01",
        html="<dl><dt>Name</dt><dd>Alice</dd><dt>Email</dt><dd>alice@example.com</dd><dt>Phone</dt><dd>555-9876</dd></dl>",
        ground_truth=KeyValueGroundTruth(
            case_id="html_dl_3fields_01",
            document_id="html_dl_3fields_01",
            source_format="html",
            detection_path="native_html_dl",
            is_key_value_group=True,
        ),
    ))

    # H3. DL with 4 fields
    cases.append(HtmlCase(
        case_id="html_dl_4fields_01",
        html="<dl><dt>Title</dt><dd>Project Alpha</dd><dt>Author</dt><dd>Bob Smith</dd><dt>Version</dt><dd>1.0</dd><dt>Status</dt><dd>Draft</dd></dl>",
        ground_truth=KeyValueGroundTruth(
            case_id="html_dl_4fields_01",
            document_id="html_dl_4fields_01",
            source_format="html",
            detection_path="native_html_dl",
            is_key_value_group=True,
        ),
    ))

    # H4. DL with link inside dd
    cases.append(HtmlCase(
        case_id="html_dl_link_01",
        html='<dl><dt>Website</dt><dd><a href="https://example.com">example.com</a></dd><dt>Contact</dt><dd>info@example.com</dd></dl>',
        ground_truth=KeyValueGroundTruth(
            case_id="html_dl_link_01",
            document_id="html_dl_link_01",
            source_format="html",
            detection_path="native_html_dl",
            is_key_value_group=True,
        ),
    ))

    # H5. DL with unicode
    cases.append(HtmlCase(
        case_id="html_dl_unicode_01",
        html="<dl><dt>Organisation</dt><dd>Académie française</dd><dt>City</dt><dd>Münich</dd><dt>Contact</dt><dd>info@acad.fr</dd></dl>",
        ground_truth=KeyValueGroundTruth(
            case_id="html_dl_unicode_01",
            document_id="html_dl_unicode_01",
            source_format="html",
            detection_path="native_html_dl",
            is_key_value_group=True,
        ),
    ))

    # H6. DL 5 fields (schedule-like)
    cases.append(HtmlCase(
        case_id="html_dl_schedule_01",
        html="<dl><dt>Monday</dt><dd>9:00 AM</dd><dt>Tuesday</dt><dd>9:00 AM</dd><dt>Wednesday</dt><dd>10:00 AM</dd><dt>Thursday</dt><dd>9:00 AM</dd><dt>Friday</dt><dd>Closed</dd></dl>",
        ground_truth=KeyValueGroundTruth(
            case_id="html_dl_schedule_01",
            document_id="html_dl_schedule_01",
            source_format="html",
            detection_path="native_html_dl",
            is_key_value_group=True,
            group_type=KeyValueGroupType.SCHEDULE,
        ),
    ))

    # H7. DL with bold dt
    cases.append(HtmlCase(
        case_id="html_dl_bold_dt_01",
        html="<dl><dt><strong>Model</strong></dt><dd>XR-500</dd><dt><strong>Serial</strong></dt><dd>SN-98765</dd><dt><strong>Year</strong></dt><dd>2022</dd></dl>",
        ground_truth=KeyValueGroundTruth(
            case_id="html_dl_bold_dt_01",
            document_id="html_dl_bold_dt_01",
            source_format="html",
            detection_path="native_html_dl",
            is_key_value_group=True,
        ),
    ))

    # H8. DL inside div
    cases.append(HtmlCase(
        case_id="html_dl_in_div_01",
        html="<div class='metadata'><dl><dt>Project</dt><dd>AksharaMD</dd><dt>Owner</dt><dd>Kalyan</dd><dt>Start</dt><dd>2024-01-01</dd></dl></div>",
        ground_truth=KeyValueGroundTruth(
            case_id="html_dl_in_div_01",
            document_id="html_dl_in_div_01",
            source_format="html",
            detection_path="native_html_dl",
            is_key_value_group=True,
        ),
    ))

    # H9. DL with numeric values
    cases.append(HtmlCase(
        case_id="html_dl_numeric_01",
        html="<dl><dt>Cores</dt><dd>8</dd><dt>RAM</dt><dd>32GB</dd><dt>Storage</dt><dd>512GB</dd></dl>",
        ground_truth=KeyValueGroundTruth(
            case_id="html_dl_numeric_01",
            document_id="html_dl_numeric_01",
            source_format="html",
            detection_path="native_html_dl",
            is_key_value_group=True,
        ),
    ))

    # H10. DL with em inside dd
    cases.append(HtmlCase(
        case_id="html_dl_em_01",
        html="<dl><dt>Status</dt><dd><em>Active</em></dd><dt>Priority</dt><dd><em>High</em></dd><dt>Due</dt><dd>2024-09-30</dd></dl>",
        ground_truth=KeyValueGroundTruth(
            case_id="html_dl_em_01",
            document_id="html_dl_em_01",
            source_format="html",
            detection_path="native_html_dl",
            is_key_value_group=True,
        ),
    ))

    return cases


def _html_negative_cases() -> list[HtmlCase]:
    cases = []

    # HN1. Empty DL (0 pairs) — no KV block expected
    cases.append(HtmlCase(
        case_id="html_dl_empty_01",
        html="<dl></dl>",
        ground_truth=KeyValueGroundTruth(
            case_id="html_dl_empty_01",
            document_id="html_dl_empty_01",
            source_format="html",
            detection_path="native_html_dl",
            is_key_value_group=False,
            negative_reason="empty_dl",
        ),
    ))

    # HN2. No DL at all
    cases.append(HtmlCase(
        case_id="html_no_dl_01",
        html="<p>This is a plain paragraph with no structured data.</p>",
        ground_truth=KeyValueGroundTruth(
            case_id="html_no_dl_01",
            document_id="html_no_dl_01",
            source_format="html",
            detection_path="native_html_dl",
            is_key_value_group=False,
            negative_reason="no_dl_element",
        ),
    ))

    # HN3. DT without matching DD (only 1 dt+dd pair — native path should still emit)
    # Actually the native path emits even for single pairs; mark as positive
    cases.append(HtmlCase(
        case_id="html_dl_single_pair_01",
        html="<dl><dt>Email</dt><dd>test@example.com</dd></dl>",
        ground_truth=KeyValueGroundTruth(
            case_id="html_dl_single_pair_01",
            document_id="html_dl_single_pair_01",
            source_format="html",
            detection_path="native_html_dl",
            is_key_value_group=True,
            notes="Native path emits even single pairs (extracted confidence)",
        ),
    ))

    # HN4. Table (not DL)
    cases.append(HtmlCase(
        case_id="html_table_not_dl_01",
        html="<table><tr><th>Name</th><th>Value</th></tr><tr><td>foo</td><td>bar</td></tr></table>",
        ground_truth=KeyValueGroundTruth(
            case_id="html_table_not_dl_01",
            document_id="html_table_not_dl_01",
            source_format="html",
            detection_path="native_html_dl",
            is_key_value_group=False,
            negative_reason="no_dl_element",
        ),
    ))

    # HN5. Unordered list (not DL)
    cases.append(HtmlCase(
        case_id="html_ul_not_dl_01",
        html="<ul><li>Item one</li><li>Item two</li><li>Item three</li></ul>",
        ground_truth=KeyValueGroundTruth(
            case_id="html_ul_not_dl_01",
            document_id="html_ul_not_dl_01",
            source_format="html",
            detection_path="native_html_dl",
            is_key_value_group=False,
            negative_reason="no_dl_element",
        ),
    ))

    return cases


# ── XLSX cases (using openpyxl in-memory) ────────────────────────────────────

def _build_xlsx_file(rows: list[list], tmpdir: str) -> str:
    """Build an XLSX file with given rows and return the file path."""
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    for row in rows:
        ws.append(row)
    path = os.path.join(tmpdir, f"test_{abs(hash(str(rows))) % 100000}.xlsx")
    wb.save(path)
    return path


def _xlsx_cases(tmpdir: str) -> list[XlsxCase]:
    cases = []

    # X1. 2-column metadata (3+ rows) → should be KV
    path1 = _build_xlsx_file([
        ["Title", "Annual Report 2024"],
        ["Author", "Finance Team"],
        ["Version", "1.0"],
        ["Date", "2024-07-01"],
    ], tmpdir)
    cases.append(XlsxCase(
        case_id="xlsx_metadata_01",
        xlsx_path=path1,
        ground_truth=KeyValueGroundTruth(
            case_id="xlsx_metadata_01",
            document_id="xlsx_metadata_01",
            source_format="xlsx",
            detection_path="native_xlsx_kv",
            is_key_value_group=True,
        ),
    ))

    # X2. 2-column with "Name"/"Value" first row → should remain table (generic header)
    path2 = _build_xlsx_file([
        ["Name", "Value"],
        ["Revenue", "1000000"],
        ["Cost", "750000"],
        ["Profit", "250000"],
    ], tmpdir)
    cases.append(XlsxCase(
        case_id="xlsx_generic_header_01",
        xlsx_path=path2,
        ground_truth=KeyValueGroundTruth(
            case_id="xlsx_generic_header_01",
            document_id="xlsx_generic_header_01",
            source_format="xlsx",
            detection_path="native_xlsx_kv",
            is_key_value_group=False,
            negative_reason="generic_column_header",
        ),
    ))

    # X3. 4-column sheet → should remain table
    path3 = _build_xlsx_file([
        ["ID", "Name", "Dept", "Salary"],
        ["1", "Alice", "Eng", "90000"],
        ["2", "Bob", "HR", "70000"],
        ["3", "Carol", "Finance", "80000"],
    ], tmpdir)
    cases.append(XlsxCase(
        case_id="xlsx_4col_01",
        xlsx_path=path3,
        ground_truth=KeyValueGroundTruth(
            case_id="xlsx_4col_01",
            document_id="xlsx_4col_01",
            source_format="xlsx",
            detection_path="native_xlsx_kv",
            is_key_value_group=False,
            negative_reason="not_2_columns",
        ),
    ))

    # X4. 2-column with repeated first-column values → should remain table
    path4 = _build_xlsx_file([
        ["Category", "Amount"],
        ["Revenue", "500"],
        ["Revenue", "600"],
        ["Cost", "300"],
        ["Cost", "400"],
    ], tmpdir)
    cases.append(XlsxCase(
        case_id="xlsx_repeated_keys_01",
        xlsx_path=path4,
        ground_truth=KeyValueGroundTruth(
            case_id="xlsx_repeated_keys_01",
            document_id="xlsx_repeated_keys_01",
            source_format="xlsx",
            detection_path="native_xlsx_kv",
            is_key_value_group=False,
            negative_reason="repeated_first_column",
        ),
    ))

    # X5. 2-column with >20 rows → should remain table (current threshold)
    big_rows = [[f"Field{i}", f"Value{i}"] for i in range(1, 25)]
    path5 = _build_xlsx_file(big_rows, tmpdir)
    cases.append(XlsxCase(
        case_id="xlsx_too_many_rows_01",
        xlsx_path=path5,
        ground_truth=KeyValueGroundTruth(
            case_id="xlsx_too_many_rows_01",
            document_id="xlsx_too_many_rows_01",
            source_format="xlsx",
            detection_path="native_xlsx_kv",
            is_key_value_group=False,
            negative_reason="too_many_rows",
        ),
    ))

    # X6. 2-column, 3 rows, short labels → should be KV
    path6 = _build_xlsx_file([
        ["Host", "localhost"],
        ["Port", "5432"],
        ["Database", "mydb"],
    ], tmpdir)
    cases.append(XlsxCase(
        case_id="xlsx_server_config_01",
        xlsx_path=path6,
        ground_truth=KeyValueGroundTruth(
            case_id="xlsx_server_config_01",
            document_id="xlsx_server_config_01",
            source_format="xlsx",
            detection_path="native_xlsx_kv",
            is_key_value_group=True,
        ),
    ))

    # X7. 2-column contact-like
    path7 = _build_xlsx_file([
        ["First Name", "Alice"],
        ["Last Name", "Smith"],
        ["Email", "alice@example.com"],
        ["Phone", "555-1234"],
    ], tmpdir)
    cases.append(XlsxCase(
        case_id="xlsx_contact_01",
        xlsx_path=path7,
        ground_truth=KeyValueGroundTruth(
            case_id="xlsx_contact_01",
            document_id="xlsx_contact_01",
            source_format="xlsx",
            detection_path="native_xlsx_kv",
            is_key_value_group=True,
        ),
    ))

    # X8. 2-column spec sheet → should be KV
    path8 = _build_xlsx_file([
        ["Voltage", "12V"],
        ["Current", "2A"],
        ["Power", "24W"],
        ["Weight", "1.5kg"],
    ], tmpdir)
    cases.append(XlsxCase(
        case_id="xlsx_spec_01",
        xlsx_path=path8,
        ground_truth=KeyValueGroundTruth(
            case_id="xlsx_spec_01",
            document_id="xlsx_spec_01",
            source_format="xlsx",
            detection_path="native_xlsx_kv",
            is_key_value_group=True,
        ),
    ))

    # X9. 2-column with "key"/"value" first row → should remain table (generic header)
    path9 = _build_xlsx_file([
        ["key", "value"],
        ["alpha", "1"],
        ["beta", "2"],
        ["gamma", "3"],
    ], tmpdir)
    cases.append(XlsxCase(
        case_id="xlsx_key_value_header_01",
        xlsx_path=path9,
        ground_truth=KeyValueGroundTruth(
            case_id="xlsx_key_value_header_01",
            document_id="xlsx_key_value_header_01",
            source_format="xlsx",
            detection_path="native_xlsx_kv",
            is_key_value_group=False,
            negative_reason="generic_column_header",
        ),
    ))

    # X10. 2-column, 5 rows, product metadata → should be KV
    path10 = _build_xlsx_file([
        ["Product", "Widget Pro"],
        ["SKU", "WGT-001"],
        ["Category", "Hardware"],
        ["Price", "29.99"],
        ["Stock", "150"],
    ], tmpdir)
    cases.append(XlsxCase(
        case_id="xlsx_product_01",
        xlsx_path=path10,
        ground_truth=KeyValueGroundTruth(
            case_id="xlsx_product_01",
            document_id="xlsx_product_01",
            source_format="xlsx",
            detection_path="native_xlsx_kv",
            is_key_value_group=True,
        ),
    ))

    return cases


# ── Abergowrie regression fixture ────────────────────────────────────────────

def abergowrie_case() -> TextCase:
    """The locked Abergowrie regression fixture — two service records."""
    text = (
        "Location: Abergowrie\n"
        "Day: Saturday\n"
        "Time: 7:00 PM\n"
        "Location: Abergowrie\n"
        "Day: Sunday\n"
        "Time: 9:00 AM"
    )
    gt = KeyValueGroundTruth(
        case_id="abergowrie_schedule",
        document_id="abergowrie_schedule",
        source_format="text",
        detection_path="heuristic_inline",
        is_key_value_group=True,
        group_type=KeyValueGroupType.SCHEDULE,
        title="Sunday Masses 29/30 September",
        records=[
            [
                GroundTruthEntry(key="Location", value="Abergowrie", record_number=0),
                GroundTruthEntry(key="Day", value="Saturday", record_number=0),
                GroundTruthEntry(key="Time", value="7:00 PM",
                                 expected_value_type=KeyValueValueType.TIME, record_number=0),
            ],
            [
                GroundTruthEntry(key="Location", value="Abergowrie", record_number=1),
                GroundTruthEntry(key="Day", value="Sunday", record_number=1),
                GroundTruthEntry(key="Time", value="9:00 AM",
                                 expected_value_type=KeyValueValueType.TIME, record_number=1),
            ],
        ],
        notes="Locked regression: Saturday 7:00 PM, Sunday 9:00 AM — must not conflate",
    )
    return TextCase(case_id="abergowrie_schedule", text=text, ground_truth=gt)


# ── Repeated record cases ─────────────────────────────────────────────────────

def _repeated_record_cases() -> list[TextCase]:
    cases = []

    # R1. Abergowrie (2 services) — duplicate of abergowrie_case but in repeated section
    ab = abergowrie_case()
    cases.append(ab)

    # R2. Two contact records (Alice/Bob)
    cases.append(TextCase(
        case_id="multi_contact_01",
        text="Name: Alice\nEmail: alice@example.com\nName: Bob\nEmail: bob@example.com",
        ground_truth=KeyValueGroundTruth(
            case_id="multi_contact_01",
            document_id="multi_contact_01",
            source_format="text",
            detection_path="heuristic_inline",
            is_key_value_group=True,
            group_type=KeyValueGroupType.CONTACT,
            records=[
                [GroundTruthEntry(key="Name", value="Alice", record_number=0),
                 GroundTruthEntry(key="Email", value="alice@example.com", record_number=0)],
                [GroundTruthEntry(key="Name", value="Bob", record_number=1),
                 GroundTruthEntry(key="Email", value="bob@example.com", record_number=1)],
            ],
        ),
    ))

    # R3. Three event records (same location, different dates)
    cases.append(TextCase(
        case_id="multi_event_01",
        text="Date: 01/07/2024\nTime: 6:00 PM\nDate: 08/07/2024\nTime: 6:00 PM\nDate: 15/07/2024\nTime: 6:00 PM",
        ground_truth=KeyValueGroundTruth(
            case_id="multi_event_01",
            document_id="multi_event_01",
            source_format="text",
            detection_path="heuristic_inline",
            is_key_value_group=True,
            group_type=KeyValueGroupType.SCHEDULE,
        ),
    ))

    # R4. Multi-week schedule (Mon-Fri pattern repeated)
    cases.append(TextCase(
        case_id="multi_week_schedule_01",
        text="Monday: 9:00 AM\nFriday: 5:00 PM\nMonday: 9:00 AM\nFriday: 5:00 PM",
        ground_truth=KeyValueGroundTruth(
            case_id="multi_week_schedule_01",
            document_id="multi_week_schedule_01",
            source_format="text",
            detection_path="heuristic_inline",
            is_key_value_group=True,
            group_type=KeyValueGroupType.SCHEDULE,
        ),
    ))

    # R5. Name/Role repeated
    cases.append(TextCase(
        case_id="multi_person_01",
        text="Name: Carol\nRole: Manager\nName: David\nRole: Engineer",
        ground_truth=KeyValueGroundTruth(
            case_id="multi_person_01",
            document_id="multi_person_01",
            source_format="text",
            detection_path="heuristic_inline",
            is_key_value_group=True,
            group_type=KeyValueGroupType.UNKNOWN,
        ),
    ))

    # R6. Product repeated pattern
    cases.append(TextCase(
        case_id="multi_product_01",
        text="Model: A100\nPrice: $99\nModel: B200\nPrice: $149",
        ground_truth=KeyValueGroundTruth(
            case_id="multi_product_01",
            document_id="multi_product_01",
            source_format="text",
            detection_path="heuristic_inline",
            is_key_value_group=True,
            group_type=KeyValueGroupType.SPECIFICATION,
        ),
    ))

    # R7. Location/Time pair repeated 3 times
    cases.append(TextCase(
        case_id="multi_location_time_01",
        text="Location: Room A\nTime: 9:00 AM\nLocation: Room B\nTime: 2:00 PM\nLocation: Room C\nTime: 4:00 PM",
        ground_truth=KeyValueGroundTruth(
            case_id="multi_location_time_01",
            document_id="multi_location_time_01",
            source_format="text",
            detection_path="heuristic_inline",
            is_key_value_group=True,
            group_type=KeyValueGroupType.SCHEDULE,
        ),
    ))

    # R8. Employee repeated records
    cases.append(TextCase(
        case_id="multi_employee_01",
        text="Employee: Eve\nDept: Sales\nEmployee: Frank\nDept: Marketing",
        ground_truth=KeyValueGroundTruth(
            case_id="multi_employee_01",
            document_id="multi_employee_01",
            source_format="text",
            detection_path="heuristic_inline",
            is_key_value_group=True,
            group_type=KeyValueGroupType.UNKNOWN,
        ),
    ))

    # R9. Course/Time repeated
    cases.append(TextCase(
        case_id="multi_course_01",
        text="Course: Math\nTime: 8:00 AM\nCourse: Science\nTime: 10:00 AM\nCourse: English\nTime: 1:00 PM",
        ground_truth=KeyValueGroundTruth(
            case_id="multi_course_01",
            document_id="multi_course_01",
            source_format="text",
            detection_path="heuristic_inline",
            is_key_value_group=True,
            group_type=KeyValueGroupType.SCHEDULE,
        ),
    ))

    # R10. Service/Location/Time 3-field repeated records
    cases.append(TextCase(
        case_id="multi_service_01",
        text="Service: Breakfast\nLocation: Dining Hall\nTime: 7:30 AM\nService: Lunch\nLocation: Dining Hall\nTime: 12:00 PM",
        ground_truth=KeyValueGroundTruth(
            case_id="multi_service_01",
            document_id="multi_service_01",
            source_format="text",
            detection_path="heuristic_inline",
            is_key_value_group=True,
            group_type=KeyValueGroupType.SCHEDULE,
        ),
    ))

    return cases


# ── Table vs KV conflict cases ───────────────────────────────────────────────

def _table_vs_kv_cases() -> list[TextCase]:
    cases = []

    # T1. Project metadata (KV: short labels, no repeated first-col)
    cases.append(TextCase(
        case_id="project_kv_vs_table_01",
        text="Project: Horizon\nLead: Alice\nBudget: $50,000\nTimeline: Q3 2024",
        ground_truth=KeyValueGroundTruth(
            case_id="project_kv_vs_table_01",
            document_id="project_kv_vs_table_01",
            source_format="text",
            detection_path="heuristic_inline",
            is_key_value_group=True,
            group_type=KeyValueGroupType.METADATA,
        ),
    ))

    # T2. Spec-like short block (KV)
    cases.append(TextCase(
        case_id="spec_kv_01",
        text="Manufacturer: Sony\nModel: WH-1000XM5\nWeight: 250g\nBattery: 30h",
        ground_truth=KeyValueGroundTruth(
            case_id="spec_kv_01",
            document_id="spec_kv_01",
            source_format="text",
            detection_path="heuristic_inline",
            is_key_value_group=True,
            group_type=KeyValueGroupType.SPECIFICATION,
        ),
    ))

    # T3. Revenue data (but short block, looks like KV)
    cases.append(TextCase(
        case_id="revenue_short_01",
        text="Q1: $250,000\nQ2: $300,000\nQ3: $275,000\nQ4: $320,000",
        ground_truth=KeyValueGroundTruth(
            case_id="revenue_short_01",
            document_id="revenue_short_01",
            source_format="text",
            detection_path="heuristic_inline",
            is_key_value_group=True,
            group_type=KeyValueGroupType.SPECIFICATION,
            notes="Technically a KV by heuristic — short labels, short values",
        ),
    ))

    # T4. Config block
    cases.append(TextCase(
        case_id="config_kv_01",
        text="Debug: false\nLogLevel: INFO\nMaxConnections: 100\nTimeout: 30s",
        ground_truth=KeyValueGroundTruth(
            case_id="config_kv_01",
            document_id="config_kv_01",
            source_format="text",
            detection_path="heuristic_inline",
            is_key_value_group=True,
            group_type=KeyValueGroupType.SPECIFICATION,
        ),
    ))

    # T5. Enrollment data (KV: no repeat, short)
    cases.append(TextCase(
        case_id="enrollment_kv_01",
        text="Program: Computer Science\nYear: 2024\nStudents: 320\nFaculty: 45",
        ground_truth=KeyValueGroundTruth(
            case_id="enrollment_kv_01",
            document_id="enrollment_kv_01",
            source_format="text",
            detection_path="heuristic_inline",
            is_key_value_group=True,
            group_type=KeyValueGroupType.SPECIFICATION,
        ),
    ))

    # T6. Contract summary
    cases.append(TextCase(
        case_id="contract_kv_01",
        text="Party A: Acme Corp\nParty B: Beta Ltd\nValue: $120,000\nDuration: 12 months",
        ground_truth=KeyValueGroundTruth(
            case_id="contract_kv_01",
            document_id="contract_kv_01",
            source_format="text",
            detection_path="heuristic_inline",
            is_key_value_group=True,
            group_type=KeyValueGroupType.FORM,
        ),
    ))

    # T7. App version metadata
    cases.append(TextCase(
        case_id="app_version_kv_01",
        text="App: MyApp\nVersion: 3.2.1\nBuild: 20240701\nPlatform: iOS",
        ground_truth=KeyValueGroundTruth(
            case_id="app_version_kv_01",
            document_id="app_version_kv_01",
            source_format="text",
            detection_path="heuristic_inline",
            is_key_value_group=True,
            group_type=KeyValueGroupType.METADATA,
        ),
    ))

    # T8. Long prose (negative — not KV)
    cases.append(TextCase(
        case_id="prose_table_conflict_01",
        text="The revenue figures for this quarter exceeded expectations.\nThe cost reduction measures are working as planned.\nFurther analysis required for Q4 projections.",
        ground_truth=KeyValueGroundTruth(
            case_id="prose_table_conflict_01",
            document_id="prose_table_conflict_01",
            source_format="text",
            detection_path="negative_control",
            is_key_value_group=False,
            negative_reason="no_colon_pattern",
        ),
    ))

    # T9. Lab results (KV)
    cases.append(TextCase(
        case_id="lab_results_01",
        text="Sample ID: LAB-001\nTest: Glucose\nResult: 5.2 mmol/L\nStatus: Normal",
        ground_truth=KeyValueGroundTruth(
            case_id="lab_results_01",
            document_id="lab_results_01",
            source_format="text",
            detection_path="heuristic_inline",
            is_key_value_group=True,
            group_type=KeyValueGroupType.FORM,
        ),
    ))

    # T10. Hotel booking
    cases.append(TextCase(
        case_id="hotel_booking_01",
        text="Hotel: Grand Palace\nRoom: 205\nCheck-in: 15/07/2024\nCheck-out: 18/07/2024",
        ground_truth=KeyValueGroundTruth(
            case_id="hotel_booking_01",
            document_id="hotel_booking_01",
            source_format="text",
            detection_path="heuristic_inline",
            is_key_value_group=True,
            group_type=KeyValueGroupType.EVENT,
        ),
    ))

    return cases


# ── DOCX properties stub ──────────────────────────────────────────────────────

@dataclass
class DocxPropertiesCase:
    """Stub for DOCX properties — records expected extraction via native docx path."""
    case_id: str
    expected_fields: dict[str, str]
    ground_truth: KeyValueGroundTruth


def _docx_properties_cases() -> list[DocxPropertiesCase]:
    return [
        DocxPropertiesCase(
            case_id="docx_properties_01",
            expected_fields={"Author": "Alice Smith", "Title": "Annual Report"},
            ground_truth=KeyValueGroundTruth(
                case_id="docx_properties_01",
                document_id="docx_properties_01",
                source_format="docx",
                detection_path="native_docx_props",
                is_key_value_group=True,
                notes="DOCX properties stub — not evaluated through compile path in dev corpus",
            ),
        ),
    ]


# ── AdjacentCase dataclass ────────────────────────────────────────────────────

@dataclass
class AdjacentCase:
    case_id: str
    blocks: list  # list of Block objects (pre-split paragraphs)
    ground_truth: KeyValueGroundTruth


def _make_para(content: str, index: int, page: int = 1):
    """Create a minimal PARAGRAPH Block for adjacent-block test cases."""
    from aksharamd.models.block import Block, BlockType
    return Block(type=BlockType.PARAGRAPH, content=content, index=index, page=page)


# ── Adjacent block cases ──────────────────────────────────────────────────────

def _adjacent_block_cases() -> list[AdjacentCase]:
    """Adjacent-block cases: each KV pair is in a separate paragraph block.

    Tests the adjacent-block promoter path (_pass_adjacent). The promoter
    joins adjacent blocks with newlines and runs the inline detector, so
    each block must contain a full 'Key: Value' line for detection to succeed.

    Note: the current promoter does NOT handle key-only / value-only block
    alternation (e.g., 'Email:' + 'alice@...' as separate blocks), because
    the inline detector requires 'Key: Value' on a single line.
    """
    cases = []

    # ADJ_01: 4 blocks, each with full 'Key: Value' — email + phone + name + city
    # The adjacent promoter joins them into one text and runs inline detection.
    cases.append(AdjacentCase(
        case_id="adj_email_phone_01",
        blocks=[
            _make_para("Email: alice@example.com", 0),
            _make_para("Phone: 555-0100", 1),
            _make_para("Name: Alice Smith", 2),
            _make_para("City: Sydney", 3),
        ],
        ground_truth=KeyValueGroundTruth(
            case_id="adj_email_phone_01",
            document_id="adj_email_phone_01",
            source_format="text",
            detection_path="heuristic_adjacent",
            is_key_value_group=True,
            notes="4 blocks: each has Key: Value format; adjacent promoter joins and detects",
        ),
    ))

    # ADJ_02: 6 blocks — name + department + title
    cases.append(AdjacentCase(
        case_id="adj_name_dept_title_01",
        blocks=[
            _make_para("Name: Alice Smith", 0),
            _make_para("Department: Engineering", 1),
            _make_para("Title: Senior Engineer", 2),
            _make_para("Start: 2022-01-15", 3),
            _make_para("Location: Building B", 4),
            _make_para("Badge: ENG-4521", 5),
        ],
        ground_truth=KeyValueGroundTruth(
            case_id="adj_name_dept_title_01",
            document_id="adj_name_dept_title_01",
            source_format="text",
            detection_path="heuristic_adjacent",
            is_key_value_group=True,
            notes="6 blocks: HR record fields each in a separate block",
        ),
    ))

    # ADJ_03: 4 blocks — generic server config labels.
    #
    # Under kv_promoter/v2 this case has zero strongly-typed values and
    # zero recognised schema fields (Host/Port/Database/SSL are not in the
    # standard contact/schedule/spec schemas). The v2 classifier therefore
    # rejects it for "insufficient_positive_evidence". This is the correct
    # behaviour: without schema signals or typed values, the promoter
    # cannot distinguish arbitrary "Label: value" pairs from prose.
    cases.append(AdjacentCase(
        case_id="adj_server_config_01",
        blocks=[
            _make_para("Host: localhost", 0),
            _make_para("Port: 5432", 1),
            _make_para("Database: prod", 2),
            _make_para("SSL: enabled", 3),
        ],
        ground_truth=KeyValueGroundTruth(
            case_id="adj_server_config_01",
            document_id="adj_server_config_01",
            source_format="text",
            detection_path="heuristic_adjacent",
            is_key_value_group=False,
            negative_reason="insufficient_positive_evidence",
            notes=(
                "kv_promoter/v2: generic labels with no schema/typed evidence "
                "-> reject. Relabelled from v1 (which promoted)."
            ),
        ),
    ))

    # ADJ_04: 6 blocks — event metadata
    cases.append(AdjacentCase(
        case_id="adj_event_01",
        blocks=[
            _make_para("Location: Room A", 0),
            _make_para("Date: 15/07/2024", 1),
            _make_para("Time: 2:00 PM", 2),
            _make_para("Organizer: Alice", 3),
            _make_para("Room: Conference B", 4),
            _make_para("Capacity: 20", 5),
        ],
        ground_truth=KeyValueGroundTruth(
            case_id="adj_event_01",
            document_id="adj_event_01",
            source_format="text",
            detection_path="heuristic_adjacent",
            is_key_value_group=True,
            notes="6 blocks: event metadata each in a separate block",
        ),
    ))

    # ADJ_NEG_01: 2 blocks — below minimum run length of 4
    # Combined text: "Conclusion:\nThe experiment did not support the hypothesis."
    # Even if promoted, only 1 candidate (if any)
    cases.append(AdjacentCase(
        case_id="adj_neg_long_value_01",
        blocks=[
            _make_para("Conclusion: The experiment did not support the hypothesis.", 0),
            _make_para("Summary: Further study is needed before drawing conclusions.", 1),
        ],
        ground_truth=KeyValueGroundTruth(
            case_id="adj_neg_long_value_01",
            document_id="adj_neg_long_value_01",
            source_format="text",
            detection_path="heuristic_adjacent",
            is_key_value_group=False,
            negative_reason="value_too_long",
            notes="2 blocks with long values — values exceed 80 chars; not promoted",
        ),
    ))

    # ADJ_NEG_02: 2 blocks — rhetorical labels (only 2 blocks, below min-4 run)
    cases.append(AdjacentCase(
        case_id="adj_neg_rhetorical_01",
        blocks=[
            _make_para("Introduction: This section covers the background of the project.", 0),
            _make_para("Summary: This document outlines the key steps for onboarding.", 1),
        ],
        ground_truth=KeyValueGroundTruth(
            case_id="adj_neg_rhetorical_01",
            document_id="adj_neg_rhetorical_01",
            source_format="text",
            detection_path="heuristic_adjacent",
            is_key_value_group=False,
            negative_reason="rhetorical_label",
            notes="Rhetorical labels (Introduction/Summary) are in rejection set; values also too long",
        ),
    ))

    # ADJ_NEG_03: 4 blocks — no colons in blocks (not KV pattern)
    cases.append(AdjacentCase(
        case_id="adj_neg_no_colon_01",
        blocks=[
            _make_para("Executive Summary", 0),
            _make_para("Our Q3 results exceeded expectations.", 1),
            _make_para("Key Finding", 2),
            _make_para("Revenue grew 15% year-over-year.", 3),
        ],
        ground_truth=KeyValueGroundTruth(
            case_id="adj_neg_no_colon_01",
            document_id="adj_neg_no_colon_01",
            source_format="text",
            detection_path="heuristic_adjacent",
            is_key_value_group=False,
            negative_reason="no_colon_in_keys",
            notes="4 blocks but no colons — no KV candidates in combined text",
        ),
    ))

    # ── kv_promoter/v2 alternating key-only / value-only cases ────────────
    # These exercise Strategy 2 in _try_promote_adjacent_run: pairs of
    # blocks where the first is "Label:" and the second holds only the
    # value. Under v2 the promoter reconstructs virtual "Key: Value" text
    # and passes it back through the classifier.

    # ADJ_ALT_01: contact card in alternating blocks — should promote
    # (Rule A: 2 strongly-typed values — email + phone)
    cases.append(AdjacentCase(
        case_id="adj_alt_contact_01",
        blocks=[
            _make_para("Email:", 0),
            _make_para("alice@example.com", 1),
            _make_para("Phone:", 2),
            _make_para("+1-555-0100", 3),
        ],
        ground_truth=KeyValueGroundTruth(
            case_id="adj_alt_contact_01",
            document_id="adj_alt_contact_01",
            source_format="text",
            detection_path="heuristic_adjacent",
            is_key_value_group=True,
            notes="Alternating key-only/value-only blocks — Strategy 2",
        ),
    ))

    # ADJ_ALT_02: schedule alternating (Rule A + Rule B)
    cases.append(AdjacentCase(
        case_id="adj_alt_schedule_01",
        blocks=[
            _make_para("Location:", 0),
            _make_para("Abergowrie", 1),
            _make_para("Day:", 2),
            _make_para("Sunday", 3),
            _make_para("Time:", 4),
            _make_para("9:00 AM", 5),
        ],
        ground_truth=KeyValueGroundTruth(
            case_id="adj_alt_schedule_01",
            document_id="adj_alt_schedule_01",
            source_format="text",
            detection_path="heuristic_adjacent",
            is_key_value_group=True,
            notes="Alternating schedule — Strategy 2 + Rule B (schedule schema)",
        ),
    ))

    # ADJ_NEG_ALT_01: dialogue in alternating blocks — reject via exclusion
    cases.append(AdjacentCase(
        case_id="adj_neg_alt_dialogue_01",
        blocks=[
            _make_para("Alice:", 0),
            _make_para("Hello.", 1),
            _make_para("Bob:", 2),
            _make_para("Good morning.", 3),
        ],
        ground_truth=KeyValueGroundTruth(
            case_id="adj_neg_alt_dialogue_01",
            document_id="adj_neg_alt_dialogue_01",
            source_format="text",
            detection_path="heuristic_adjacent",
            is_key_value_group=False,
            negative_reason="dialogue",
            notes="Alternating blocks that form dialogue — should be rejected",
        ),
    ))

    # ADJ_NEG_ALT_02: section-label alternating — reject via exclusion
    cases.append(AdjacentCase(
        case_id="adj_neg_alt_section_01",
        blocks=[
            _make_para("Section 1:", 0),
            _make_para("Applicability", 1),
            _make_para("Section 2:", 2),
            _make_para("Definitions", 3),
        ],
        ground_truth=KeyValueGroundTruth(
            case_id="adj_neg_alt_section_01",
            document_id="adj_neg_alt_section_01",
            source_format="text",
            detection_path="heuristic_adjacent",
            is_key_value_group=False,
            negative_reason="section_label",
            notes="Alternating section labels — reject via exclusion",
        ),
    ))

    # ADJ_NEG_ALT_03: configuration in alternating blocks — reject via exclusion
    cases.append(AdjacentCase(
        case_id="adj_neg_alt_config_01",
        blocks=[
            _make_para("debug:", 0),
            _make_para("false", 1),
            _make_para("log_level:", 2),
            _make_para("info", 3),
            _make_para("port:", 4),
            _make_para("8080", 5),
        ],
        ground_truth=KeyValueGroundTruth(
            case_id="adj_neg_alt_config_01",
            document_id="adj_neg_alt_config_01",
            source_format="text",
            detection_path="heuristic_adjacent",
            is_key_value_group=False,
            negative_reason="configuration",
            notes="Alternating configuration keys — reject via exclusion",
        ),
    ))

    # ADJ_ALT_03: event alternating (Rule A: date + time strong typing)
    cases.append(AdjacentCase(
        case_id="adj_alt_event_01",
        blocks=[
            _make_para("Date:", 0),
            _make_para("15/07/2024", 1),
            _make_para("Time:", 2),
            _make_para("2:00 PM", 3),
            _make_para("Venue:", 4),
            _make_para("Grand Hall", 5),
        ],
        ground_truth=KeyValueGroundTruth(
            case_id="adj_alt_event_01",
            document_id="adj_alt_event_01",
            source_format="text",
            detection_path="heuristic_adjacent",
            is_key_value_group=True,
            notes="Alternating event fields — Rule A (date+time typed)",
        ),
    ))

    return cases


# ── Hard-negative cases ───────────────────────────────────────────────────────

def _hard_negative_cases() -> list[TextCase]:
    """Hard negatives: cases that may fool the current heuristic detector.

    These test known gaps identified by the reviewer. Labels reflect what SHOULD
    be the correct ground truth. Where the detector incorrectly promotes a case,
    that is a real FP — documented in negative_reason and notes.

    Labels are verified against detect_key_value_entries() output:
    - Cases the detector correctly rejects: is_key_value_group=False (TN)
    - Cases the detector incorrectly promotes: is_key_value_group=False (FP)
    - Cases the detector correctly promotes: is_key_value_group=True (TP)
    """
    cases = []

    # HN_A1. Legal section numbers with short values — KNOWN FP
    # detector returns: detected=True entries=3
    cases.append(TextCase(
        case_id="legal_section_short_01",
        text="Section 1: Applicability\nSection 2: Definitions\nSection 3: Obligations",
        ground_truth=KeyValueGroundTruth(
            case_id="legal_section_short_01",
            document_id="legal_section_short_01",
            source_format="text",
            detection_path="heuristic_inline",
            is_key_value_group=False,
            negative_reason="legal_clause_numbering",
            notes="Known FP: numbered section labels with short values pass heuristic (Section N is <=5 words, value is short)",
        ),
    ))

    # HN_A2. Financial footnote with colons — KNOWN FP
    # detector returns: detected=True entries=2
    cases.append(TextCase(
        case_id="financial_footnote_01",
        text="(1) Based on audited financials: FY2023\n(2) Excluding non-recurring items: Q3",
        ground_truth=KeyValueGroundTruth(
            case_id="financial_footnote_01",
            document_id="financial_footnote_01",
            source_format="text",
            detection_path="heuristic_inline",
            is_key_value_group=False,
            negative_reason="financial_footnote_with_colon",
            notes="Known FP: footnote numbering with colon passes heuristic",
        ),
    ))

    # HN_A3. Bibliographic citation (Author: Year) — KNOWN FP
    # detector returns: detected=True entries=3
    cases.append(TextCase(
        case_id="bibliographic_01",
        text="Smith: 2023\nJones: 2021\nBrown: 2019",
        ground_truth=KeyValueGroundTruth(
            case_id="bibliographic_01",
            document_id="bibliographic_01",
            source_format="text",
            detection_path="heuristic_inline",
            is_key_value_group=False,
            negative_reason="bibliographic_citation",
            notes="Known FP: Author: Year looks like KV but is a citation pattern",
        ),
    ))

    # HN_A4. Dialogue with short utterances — KNOWN FP
    # detector returns: detected=True entries=3
    cases.append(TextCase(
        case_id="dialogue_short_01",
        text="Alice: Hello there.\nBob: Good morning.\nCarol: How are you?",
        ground_truth=KeyValueGroundTruth(
            case_id="dialogue_short_01",
            document_id="dialogue_short_01",
            source_format="text",
            detection_path="heuristic_inline",
            is_key_value_group=False,
            negative_reason="dialogue_short_utterances",
            notes="Known FP: short dialogue utterances pass 80-char value limit",
        ),
    ))

    # HN_A5. API documentation (method: description) — KNOWN FP
    # detector returns: detected=True entries=3
    cases.append(TextCase(
        case_id="api_doc_01",
        text="GET: Retrieve a resource by ID\nPOST: Create a new resource\nDELETE: Remove a resource",
        ground_truth=KeyValueGroundTruth(
            case_id="api_doc_01",
            document_id="api_doc_01",
            source_format="text",
            detection_path="heuristic_inline",
            is_key_value_group=False,
            negative_reason="api_doc_method_description",
            notes="Known FP: HTTP method: description passes heuristic",
        ),
    ))

    # HN_A6. YAML-like config — KNOWN FP
    # detector returns: detected=True entries=3
    # Note: YAML config is borderline; some consider it valid KV, but
    # without parser-native structure it can't be distinguished from prose config
    cases.append(TextCase(
        case_id="yaml_like_config_01",
        text="debug: false\nlog_level: INFO\nmax_retries: 3",
        ground_truth=KeyValueGroundTruth(
            case_id="yaml_like_config_01",
            document_id="yaml_like_config_01",
            source_format="text",
            detection_path="heuristic_inline",
            is_key_value_group=False,
            negative_reason="yaml_config_format",
            notes="Known FP: YAML config passes heuristic — borderline case, no native structure",
        ),
    ))

    # HN_A7. Heading-with-descriptor (Introduction: Overview) — KNOWN FP
    # detector returns: detected=True entries=2
    cases.append(TextCase(
        case_id="heading_short_desc_01",
        text="Introduction: Overview\nBackground: History\nMethods: Approach",
        ground_truth=KeyValueGroundTruth(
            case_id="heading_short_desc_01",
            document_id="heading_short_desc_01",
            source_format="text",
            detection_path="heuristic_inline",
            is_key_value_group=False,
            negative_reason="rhetorical_label",
            notes="Known FP: heading-descriptor pairs look like KV but are section structure",
        ),
    ))

    # HN_A8. Email in prose context — KNOWN FP (via single-entry email exception)
    # detector returns: detected=True entries=1 (only first line matches; second rejected)
    # "Please send feedback to" = 4 words, under limit; email value triggers single-entry exception
    cases.append(TextCase(
        case_id="email_in_prose_01",
        text="Please send feedback to: feedback@company.com\nFor technical issues: support@company.com",
        ground_truth=KeyValueGroundTruth(
            case_id="email_in_prose_01",
            document_id="email_in_prose_01",
            source_format="text",
            detection_path="heuristic_inline",
            is_key_value_group=False,
            negative_reason="prose_with_embedded_email",
            notes="Known FP: 'Please send feedback to' (4 words) + email value triggers single-entry exception",
        ),
    ))

    # HN_A9. Flattened quarterly table (Q1: 250000) — KNOWN FP
    # detector returns: detected=True entries=4
    cases.append(TextCase(
        case_id="flattened_table_01",
        text="Q1: 250000\nQ2: 300000\nQ3: 275000\nQ4: 320000",
        ground_truth=KeyValueGroundTruth(
            case_id="flattened_table_01",
            document_id="flattened_table_01",
            source_format="text",
            detection_path="heuristic_inline",
            is_key_value_group=False,
            negative_reason="quarterly_data_table",
            notes="Known FP: Quarterly financial data looks like KV but is tabular data",
        ),
    ))

    # HN_A10. OCR-corrupted spacing (space before colon) — TN (correctly rejected)
    # detector returns: detected=False reason=no_candidate_pairs
    cases.append(TextCase(
        case_id="ocr_corrupt_01",
        text="Name :John Smith\nPhone :555-1234\nEmail :john@example.com",
        ground_truth=KeyValueGroundTruth(
            case_id="ocr_corrupt_01",
            document_id="ocr_corrupt_01",
            source_format="text",
            detection_path="heuristic_inline",
            is_key_value_group=False,
            negative_reason="ocr_corrupted_spacing",
            notes="TN: space before colon breaks ': ' separator; correctly not detected",
        ),
    ))

    # HN_A11. Transcript (Moderator: text) — KNOWN FP
    # detector returns: detected=True entries=3
    cases.append(TextCase(
        case_id="transcript_01",
        text="Moderator: Thank you all for joining.\nPanelist 1: Happy to be here today.\nPanelist 2: Excited to discuss this topic.",
        ground_truth=KeyValueGroundTruth(
            case_id="transcript_01",
            document_id="transcript_01",
            source_format="text",
            detection_path="heuristic_inline",
            is_key_value_group=False,
            negative_reason="transcript_speaker_labels",
            notes="Known FP: speaker: utterance pattern passes heuristic",
        ),
    ))

    # HN_A12. Medical report sections — KNOWN FP
    # detector returns: detected=True entries=3
    cases.append(TextCase(
        case_id="medical_finding_01",
        text="Impression: Unremarkable\nFindings: Normal\nRecommendation: Follow-up",
        ground_truth=KeyValueGroundTruth(
            case_id="medical_finding_01",
            document_id="medical_finding_01",
            source_format="text",
            detection_path="heuristic_inline",
            is_key_value_group=False,
            negative_reason="medical_section_labels",
            notes="Known FP: medical section labels with short values pass heuristic; borderline",
        ),
    ))

    # HN_A13. French contact labels — VALID KV (multilingual positive)
    # detector returns: detected=True entries=3
    cases.append(TextCase(
        case_id="multilingual_fr_01",
        text="Nom: Jean Dupont\nVille: Paris\nPays: France",
        ground_truth=KeyValueGroundTruth(
            case_id="multilingual_fr_01",
            document_id="multilingual_fr_01",
            source_format="text",
            detection_path="heuristic_inline",
            is_key_value_group=True,
            notes="French contact labels — valid KV, correctly detected",
        ),
    ))

    # HN_A14. Numbered list items — KNOWN FP
    # detector returns: detected=True entries=3
    cases.append(TextCase(
        case_id="numbered_list_01",
        text="1: First item\n2: Second item\n3: Third item",
        ground_truth=KeyValueGroundTruth(
            case_id="numbered_list_01",
            document_id="numbered_list_01",
            source_format="text",
            detection_path="heuristic_inline",
            is_key_value_group=False,
            negative_reason="numbered_list_items",
            notes="Known FP: integer labels pass heuristic but are list ordinals not field names",
        ),
    ))

    # HN_A15. Server config (host: localhost) — VALID KV
    # detector returns: detected=True entries=3
    cases.append(TextCase(
        case_id="code_config_01",
        text="host: localhost\nport: 8080\ntimeout: 30",
        ground_truth=KeyValueGroundTruth(
            case_id="code_config_01",
            document_id="code_config_01",
            source_format="text",
            detection_path="heuristic_inline",
            is_key_value_group=True,
            notes="Server config parameters — valid KV, correctly detected",
        ),
    ))

    # HN_A16. Academic definitions — KNOWN FP
    # detector returns: detected=True entries=3
    cases.append(TextCase(
        case_id="academic_def_01",
        text="Entropy: A measure of disorder\nEnthalpy: A measure of heat content\nGibbs: Free energy function",
        ground_truth=KeyValueGroundTruth(
            case_id="academic_def_01",
            document_id="academic_def_01",
            source_format="text",
            detection_path="heuristic_inline",
            is_key_value_group=False,
            negative_reason="academic_term_definition",
            notes="Known FP: scientific term: short definition passes heuristic",
        ),
    ))

    # HN_A17. Financial summary (Revenue: $1.2M) — VALID KV
    # detector returns: detected=True entries=3
    cases.append(TextCase(
        case_id="financial_statement_01",
        text="Revenue: $1.2M\nExpenses: $0.9M\nNet: $0.3M",
        ground_truth=KeyValueGroundTruth(
            case_id="financial_statement_01",
            document_id="financial_statement_01",
            source_format="text",
            detection_path="heuristic_inline",
            is_key_value_group=True,
            notes="Financial summary block — valid KV with currency-like values",
        ),
    ))

    # HN_A18. Resource URL labels — VALID KV
    # detector returns: detected=True entries=2
    cases.append(TextCase(
        case_id="url_in_prose_01",
        text="Documentation: https://docs.example.com\nRepository: https://github.com/example/repo",
        ground_truth=KeyValueGroundTruth(
            case_id="url_in_prose_01",
            document_id="url_in_prose_01",
            source_format="text",
            detection_path="heuristic_inline",
            is_key_value_group=True,
            notes="Resource URL reference block — valid KV with URL values",
        ),
    ))

    # HN_A19. Ratio values with embedded colons — VALID KV
    # detector returns: detected=True entries=3
    # first colon is the KV separator; value may contain additional colons
    cases.append(TextCase(
        case_id="unit_colon_01",
        text="Ratio: 3:1\nScale: 1:100\nMix: 2:1",
        ground_truth=KeyValueGroundTruth(
            case_id="unit_colon_01",
            document_id="unit_colon_01",
            source_format="text",
            detection_path="heuristic_inline",
            is_key_value_group=True,
            notes="Ratio values contain colons — first colon is KV separator; valid KV",
        ),
    ))

    # HN_A20. Bibliographic metadata — VALID KV
    # detector returns: detected=True entries=4
    cases.append(TextCase(
        case_id="citation_style_01",
        text="Author: Smith et al.\nYear: 2023\nJournal: Nature\nDOI: 10.1038/s41586-023",
        ground_truth=KeyValueGroundTruth(
            case_id="citation_style_01",
            document_id="citation_style_01",
            source_format="text",
            detection_path="heuristic_inline",
            is_key_value_group=True,
            notes="Bibliographic metadata — valid KV despite citation context",
        ),
    ))

    # ── kv_promoter/v2 expanded hard-negative corpus (75+ target) ─────────
    # Cases across 9 exclusion categories. Each labeled is_key_value_group
    # =False; the v2 classifier should reject via the corresponding
    # exclusion category.

    # --- DIALOGUE (10 cases) ---
    for i, (a_key, a_val, b_key, b_val) in enumerate([
        ("Alice", "Where should we meet?", "Bob", "How about 3 PM?"),
        ("Dave", "Great idea!", "Eve", "I agree completely."),
        ("Frank", "Not sure yet.", "Grace", "Let me check my calendar."),
        ("Henry", "Yes.", "Ivy", "No."),
        ("Jack", "Ready to start.", "Kate", "Same here."),
        ("Liam", "That was fun.", "Mia", "Definitely!"),
        ("Noah", "Hello?", "Olivia", "Yes hi!"),
        ("Paul", "See you tomorrow.", "Quinn", "Sounds good."),
        ("Ruby", "How are things?", "Sam", "Pretty good thanks."),
        ("Tom", "Are you coming?", "Uma", "On my way."),
    ], start=1):
        cid = f"v2_dialogue_{i:02d}"
        cases.append(TextCase(
            case_id=cid,
            text=f"{a_key}: {a_val}\n{b_key}: {b_val}",
            ground_truth=KeyValueGroundTruth(
                case_id=cid, document_id=cid, source_format="text",
                detection_path="heuristic_inline",
                is_key_value_group=False,
                negative_reason="dialogue",
                notes="v2 dialogue exclusion — proper-noun keys + sentence-ending values",
            ),
        ))

    # --- LEGAL CLAUSE (8 cases) ---
    for i, (a, b, c) in enumerate([
        ("Applicability", "Definitions", "Obligations"),
        ("Introduction", "Scope", "Term"),
        ("Payment", "Delivery", "Warranty"),
        ("Confidentiality", "Non-disclosure", "Return"),
        ("Force Majeure", "Termination", "Notices"),
        ("Governing Law", "Jurisdiction", "Amendments"),
        ("Assignment", "Waiver", "Severability"),
        ("Entire Agreement", "Counterparts", "Effective Date"),
    ], start=1):
        cid = f"v2_legal_{i:02d}"
        cases.append(TextCase(
            case_id=cid,
            text=f"Clause 1: {a}\nClause 2: {b}\nClause 3: {c}",
            ground_truth=KeyValueGroundTruth(
                case_id=cid, document_id=cid, source_format="text",
                detection_path="heuristic_inline",
                is_key_value_group=False,
                negative_reason="legal_clause",
                notes="v2 legal_clause exclusion",
            ),
        ))

    # --- ACADEMIC DEFINITION (8 cases) ---
    for i, (term, definition, term2, def2) in enumerate([
        ("Photon", "A quantum of light", "Electron", "A subatomic particle"),
        ("Enzyme", "A biological catalyst", "Substrate", "A molecule an enzyme acts on"),
        ("Mitosis", "A form of cell division", "Meiosis", "A specialised cell division"),
        ("Algorithm", "A finite procedure for solving a problem", "Data", "The raw material of computation"),
        ("Osmosis", "The diffusion of water", "Diffusion", "The spreading of particles"),
        ("Rhythm", "A pattern of sound in time", "Melody", "A sequence of musical tones"),
        ("Neuron", "A nerve cell", "Synapse", "The junction between two neurons"),
        ("Ecosystem", "A community of interacting organisms", "Habitat", "The place where an organism lives"),
    ], start=1):
        cid = f"v2_academic_{i:02d}"
        cases.append(TextCase(
            case_id=cid,
            text=f"{term}: {definition}\n{term2}: {def2}",
            ground_truth=KeyValueGroundTruth(
                case_id=cid, document_id=cid, source_format="text",
                detection_path="heuristic_inline",
                is_key_value_group=False,
                negative_reason="academic_definition",
                notes="v2 academic_definition exclusion (values start with A/An/The)",
            ),
        ))

    # --- CITATION (6 cases) ---
    for i, (a, y1, b, y2) in enumerate([
        ("Smith", "2023", "Jones", "2021"),
        ("Kim", "2020", "Lee", "2019"),
        ("Patel", "2022", "Singh", "2018"),
        ("Chen", "2024", "Wang", "2015"),
        ("Garcia", "2017", "Lopez", "2013"),
        ("Brown", "2016", "Davis", "2014"),
    ], start=1):
        cid = f"v2_citation_{i:02d}"
        cases.append(TextCase(
            case_id=cid,
            text=f"{a}: {y1}\n{b}: {y2}",
            ground_truth=KeyValueGroundTruth(
                case_id=cid, document_id=cid, source_format="text",
                detection_path="heuristic_inline",
                is_key_value_group=False,
                negative_reason="citation",
                notes="v2 citation exclusion",
            ),
        ))

    # --- CONFIGURATION (10 cases) ---
    for i, cfg in enumerate([
        "debug: false\nlog_level: info\nmax_retries: 3",
        "verbose: true\ntimeout: 30\nretries: 5",
        "cache: enabled\ncompress: true\nquota: 500",
        "ssl: disabled\ncors: false\nauth: none",
        "dry_run: true\nforce: false\nquiet: yes",
        "loglevel: warning\nformat: json\nrotation: daily",
        "enabled: true\nreadonly: false\nfoo: bar",
        "backend: postgres\nhost: localhost\nport: 5432",
        "algorithm: sha256\nkeylength: 256\nverbose: true",
        "profile: prod\nregion: us_east\nreplicas: 3",
    ], start=1):
        cid = f"v2_config_{i:02d}"
        cases.append(TextCase(
            case_id=cid,
            text=cfg,
            ground_truth=KeyValueGroundTruth(
                case_id=cid, document_id=cid, source_format="text",
                detection_path="heuristic_inline",
                is_key_value_group=False,
                negative_reason="configuration",
                notes="v2 configuration exclusion — snake_case identifier keys",
            ),
        ))

    # --- MEDICAL SECTION (6 cases) ---
    for i, txt in enumerate([
        "Impression: Unremarkable\nFindings: Normal",
        "Assessment: Stable\nPlan: Follow-up in 3 months",
        "History: Recurring cough\nDiagnosis: Chronic bronchitis",
        "Impression: Nodule detected\nRecommendation: Biopsy",
        "Findings: No abnormality\nAssessment: Healthy",
        "Complaint: Chest pain\nExamination: EKG normal",
    ], start=1):
        cid = f"v2_medical_{i:02d}"
        cases.append(TextCase(
            case_id=cid,
            text=txt,
            ground_truth=KeyValueGroundTruth(
                case_id=cid, document_id=cid, source_format="text",
                detection_path="heuristic_inline",
                is_key_value_group=False,
                negative_reason="medical_section",
                notes="v2 medical_section exclusion",
            ),
        ))

    # --- FINANCIAL FOOTNOTE (5 cases) ---
    for i, txt in enumerate([
        "(1): Based on FY23 audit\n(2): Excludes non-recurring items",
        "1.: Adjusted EBITDA\n2.: See appendix",
        "(a): Reported figures\n(b): Restated figures",
        "1.: Pro-forma basis\n2.: GAAP basis\n3.: IFRS basis",
        "(1): Comparable\n(2): Non-comparable\n(3): Foreign exchange",
    ], start=1):
        cid = f"v2_footnote_{i:02d}"
        cases.append(TextCase(
            case_id=cid,
            text=txt,
            ground_truth=KeyValueGroundTruth(
                case_id=cid, document_id=cid, source_format="text",
                detection_path="heuristic_inline",
                is_key_value_group=False,
                negative_reason="financial_footnote",
                notes="v2 financial_footnote exclusion",
            ),
        ))

    # --- NUMBERED LIST (6 cases) ---
    for i, txt in enumerate([
        "1: First step\n2: Second step\n3: Third step",
        "1: Introduction\n2: Body\n3: Conclusion",
        "1: Alpha\n2: Beta\n3: Gamma\n4: Delta",
        "1.1: Overview\n1.2: Scope",
        "1: Requirements\n2: Design\n3: Implementation\n4: Testing",
        "1: Prep\n2: Cook\n3: Serve",
    ], start=1):
        cid = f"v2_numlist_{i:02d}"
        cases.append(TextCase(
            case_id=cid,
            text=txt,
            ground_truth=KeyValueGroundTruth(
                case_id=cid, document_id=cid, source_format="text",
                detection_path="heuristic_inline",
                is_key_value_group=False,
                negative_reason="numbered_list",
                notes="v2 numbered_list exclusion — integer/dotted keys",
            ),
        ))

    # --- SECTION LABEL (8 cases) ---
    for i, txt in enumerate([
        "Section 1: Introduction\nSection 2: Methods\nSection 3: Results",
        "Article I: Preamble\nArticle II: Legislative\nArticle III: Executive",
        "Part 1: Background\nPart 2: Findings",
        "Chapter 1: Beginnings\nChapter 2: Growth\nChapter 3: Maturity",
        "Section A: Terms\nSection B: Definitions",  # this is not numbered
        "Q1: Growth\nQ2: Stability\nQ3: Expansion\nQ4: Consolidation",
        "Section 1: Applicability\nSection 2: Scope\nSection 3: Duration",
        "Article 1: Rights\nArticle 2: Duties\nArticle 3: Enforcement",
    ], start=1):
        cid = f"v2_section_{i:02d}"
        cases.append(TextCase(
            case_id=cid,
            text=txt,
            ground_truth=KeyValueGroundTruth(
                case_id=cid, document_id=cid, source_format="text",
                detection_path="heuristic_inline",
                is_key_value_group=False,
                negative_reason="section_label",
                notes="v2 section_label exclusion",
            ),
        ))

    # --- GENERAL PROSE (8 cases) — non-KV that could still fool detector ---
    for i, txt in enumerate([
        "Overview: This document explains our approach and methodology in general terms.",
        "Warning: Do not proceed without checking safety equipment first before starting.",
        "Note: All amounts are approximate.\nCaveat: Subject to change.",
        "See: attached diagram\nReference: p.12",  # short but rhetorical labels
        "Reason: unclear\nOutcome: pending review",  # rhetorical labels
        "Purpose: onboarding\nObjective: complete training",  # rhetorical labels
        "Introduction: This chapter explains the key concepts of the field of study.",
        "Context: fiscal year 2023\nBackground: revenue slowdown observed in Q3",
    ], start=1):
        cid = f"v2_prose_{i:02d}"
        cases.append(TextCase(
            case_id=cid,
            text=txt,
            ground_truth=KeyValueGroundTruth(
                case_id=cid, document_id=cid, source_format="text",
                detection_path="heuristic_inline",
                is_key_value_group=False,
                negative_reason="general_prose_or_rhetorical",
                notes="v2 general prose — should be rejected via rhetorical labels or missing evidence",
            ),
        ))

    return cases


# ── XLSX expanded cases ───────────────────────────────────────────────────────

def _xlsx_expanded_cases(tmpdir: str) -> list[XlsxCase]:
    """Five additional XLSX cases beyond the original 10."""
    import openpyxl
    cases = []

    # X11. Mixed sheet: first 3 rows are metadata, rows 4-8 are tabular data
    # -> Should NOT be KV (repeated first-col values in data section)
    wb11 = openpyxl.Workbook()
    ws11 = wb11.active
    ws11.title = "Sheet1"
    ws11.append(["Report Title", "Q3 Financial Summary"])
    ws11.append(["Author", "Finance Team"])
    ws11.append(["Date", "2024-07-01"])
    ws11.append(["Region", "Revenue"])
    ws11.append(["North", "500000"])
    ws11.append(["North", "480000"])
    ws11.append(["South", "320000"])
    ws11.append(["South", "310000"])
    path11 = os.path.join(tmpdir, "test_mixed_sheet.xlsx")
    wb11.save(path11)
    cases.append(XlsxCase(
        case_id="xlsx_mixed_sheet_01",
        xlsx_path=path11,
        ground_truth=KeyValueGroundTruth(
            case_id="xlsx_mixed_sheet_01",
            document_id="xlsx_mixed_sheet_01",
            source_format="xlsx",
            detection_path="native_xlsx_kv",
            is_key_value_group=False,
            negative_reason="repeated_first_column",
            notes="Mixed sheet: metadata header + tabular data body; repeated first-col rejects KV",
        ),
    ))

    # X12. Metric/Value two-column table — generic headers → NOT KV
    wb12 = openpyxl.Workbook()
    ws12 = wb12.active
    ws12.title = "Sheet1"
    ws12.append(["Metric", "Value"])
    ws12.append(["Revenue", "1200000"])
    ws12.append(["Cost", "900000"])
    ws12.append(["Profit", "300000"])
    path12 = os.path.join(tmpdir, "test_metric_value.xlsx")
    wb12.save(path12)
    cases.append(XlsxCase(
        case_id="xlsx_metric_value_header_01",
        xlsx_path=path12,
        ground_truth=KeyValueGroundTruth(
            case_id="xlsx_metric_value_header_01",
            document_id="xlsx_metric_value_header_01",
            source_format="xlsx",
            detection_path="native_xlsx_kv",
            is_key_value_group=False,
            negative_reason="generic_column_header",
            notes="First row 'Metric'/'Value' matches generic header names; rejected as data table",
        ),
    ))

    # X13. 2-column with blank labels (empty first-col cells) → NOT KV
    wb13 = openpyxl.Workbook()
    ws13 = wb13.active
    ws13.title = "Sheet1"
    ws13.append(["Title", "Annual Report"])
    ws13.append(["", "Subtitle line"])
    ws13.append(["Author", "Finance Team"])
    ws13.append(["Date", "2024-07-01"])
    path13 = os.path.join(tmpdir, "test_blank_labels.xlsx")
    wb13.save(path13)
    cases.append(XlsxCase(
        case_id="xlsx_blank_labels_01",
        xlsx_path=path13,
        ground_truth=KeyValueGroundTruth(
            case_id="xlsx_blank_labels_01",
            document_id="xlsx_blank_labels_01",
            source_format="xlsx",
            detection_path="native_xlsx_kv",
            is_key_value_group=True,
            notes="Blank first-col row is skipped by parser; 3 non-blank label rows (Title/Author/Date) pass _is_kv_region",
        ),
    ))

    # X14. 2-column with merged title cell + key/value rows
    # Behavior depends on how openpyxl reads merged cells.
    # The merged row is typically read as a single row with value in first cell.
    # After skip of the merged row, remaining rows 2-5 form a 4-row 2-col region.
    wb14 = openpyxl.Workbook()
    ws14 = wb14.active
    ws14.title = "Sheet1"
    ws14.merge_cells("A1:B1")
    ws14["A1"] = "Configuration Settings"
    ws14.append(["Host", "localhost"])
    ws14.append(["Port", "5432"])
    ws14.append(["Database", "prod"])
    ws14.append(["SSL", "enabled"])
    path14 = os.path.join(tmpdir, "test_merged_title.xlsx")
    wb14.save(path14)
    # After read: merged A1:B1 creates a row with ["Configuration Settings", None]
    # The spreadsheet parser reads all populated rows. The merged row has col_count=2
    # but second col is empty (slave cell). The parser may read 5 rows total.
    # "Configuration Settings" is a 2-word label, "None" (empty) value.
    # The merged row will be included and its value in col1 is empty.
    # With 5 rows total and col_count=2: passes row count check.
    # First row headers: "configuration settings" not in _KV_HEADER_NAMES.
    # First col values: "Configuration Settings" (2 words, ok), "Host", "Port", "Database", "SSL" — all unique
    # col1 of merged row: empty string — passes (empty is ok for col1)
    # Result: is_key_value_group=True (the merged cell becomes a KV entry with empty value)
    cases.append(XlsxCase(
        case_id="xlsx_merged_title_01",
        xlsx_path=path14,
        ground_truth=KeyValueGroundTruth(
            case_id="xlsx_merged_title_01",
            document_id="xlsx_merged_title_01",
            source_format="xlsx",
            detection_path="native_xlsx_kv",
            is_key_value_group=True,
            notes="Merged title row + 4 KV rows; merged cell read as label with empty value; passes _is_kv_region",
        ),
    ))

    # X15. Exactly 20 rows — boundary test (max allowed is 20) → should be KV
    rows20 = [[f"Field{i}", f"Value{i}"] for i in range(1, 21)]
    wb15 = openpyxl.Workbook()
    ws15 = wb15.active
    ws15.title = "Sheet1"
    for row in rows20:
        ws15.append(row)
    path15 = os.path.join(tmpdir, "test_20rows.xlsx")
    wb15.save(path15)
    cases.append(XlsxCase(
        case_id="xlsx_boundary_20rows_01",
        xlsx_path=path15,
        ground_truth=KeyValueGroundTruth(
            case_id="xlsx_boundary_20rows_01",
            document_id="xlsx_boundary_20rows_01",
            source_format="xlsx",
            detection_path="native_xlsx_kv",
            is_key_value_group=True,
            notes="Exactly 20 rows — at the boundary limit; should be KV (row_count <= 20)",
        ),
    ))

    return cases


# ── Main loader ───────────────────────────────────────────────────────────────

# Module-level temp dir for XLSX files — created once per import, lives for session
_XLSX_TMPDIR: str | None = None


def _get_xlsx_tmpdir() -> str:
    global _XLSX_TMPDIR
    if _XLSX_TMPDIR is None:
        _XLSX_TMPDIR = tempfile.mkdtemp(prefix="kv_eval_xlsx_")
    return _XLSX_TMPDIR


def load_dev_corpus() -> dict[str, list]:
    """Returns dict[detection_path, list[case]] for the dev corpus.

    Keys:
    - "heuristic_inline"    — TextCase (positive)
    - "negative_control"    — TextCase (negative)
    - "native_html_dl"      — HtmlCase (positive + negative)
    - "native_xlsx_kv"      — XlsxCase
    - "repeated_records"    — TextCase (multi-record positive)
    - "table_vs_kv"         — TextCase (ambiguous/positive)
    - "hard_negative"       — TextCase (hard negative/borderline cases)
    - "adjacent_block"      — AdjacentCase (adjacent-block promoter cases)
    """
    tmpdir = _get_xlsx_tmpdir()
    return {
        "heuristic_inline": _tp_inline_cases(),
        "negative_control": _negative_control_cases(),
        "native_html_dl": _html_positive_cases() + _html_negative_cases(),
        "native_xlsx_kv": _xlsx_cases(tmpdir) + _xlsx_expanded_cases(tmpdir),
        "repeated_records": _repeated_record_cases(),
        "table_vs_kv": _table_vs_kv_cases(),
        "hard_negative": _hard_negative_cases(),
        "adjacent_block": _adjacent_block_cases(),
    }
